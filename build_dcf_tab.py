# comment to trigger build

"""
build_dcf_tab.py
================
Builds the full DCF tab in the NVIDIA analyst model Excel file.
Includes: Revenue build by segment, operating margin drivers,
UFCF build, discounting, valuation, equity bridge, NWC detail, and memo items.

Usage:
    cd Final_Project
    python build_dcf_tab.py

This is the FINAL version that includes:
- 5 segment revenue builds with Y/Y growth rates
- Gross margin, R&D, SG&A margin drivers (with bps improvements)
- Full UFCF → DCF → equity bridge → implied price
- Color coding: Blue = hardcoded inputs, Green = cross-tab refs, Black = local formulas
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

wb = openpyxl.load_workbook('NVIDIA NVDA US.xlsx')

if 'DCF' in wb.sheetnames:
    idx = wb.sheetnames.index('DCF')
    del wb['DCF']
else:
    idx = 1
ws = wb.create_sheet('DCF', idx)

# ── Column mappings ──────────────────────────────────────────
MC = {  # Model tab columns
    'FY2025':'BS','FY2026':'BX','FY2027':'CC',
    'FY2028':'CD','FY2029':'CE','FY2030':'CF',
}
DC = {  # DCF tab columns
    'FY2025':'F','FY2026':'G','FY2027':'H',
    'FY2028':'I','FY2029':'J','FY2030':'K',
}
ALL = ['FY2025','FY2026','FY2027','FY2028','FY2029','FY2030']
PROJ = ['FY2026','FY2027','FY2028','FY2029','FY2030']

# ── Colors & Fonts ───────────────────────────────────────────
BLUE   = '0000FF'
GREEN  = '007F00'
BLACK  = '000000'

def mkfont(color=BLACK, bold=False, size=None, italic=False, underline=None):
    kw = {'color': color, 'bold': bold}
    if size: kw['size'] = size
    if italic: kw['italic'] = italic
    if underline: kw['underline'] = underline
    return Font(**kw)

f_title    = mkfont(BLACK, True, 14)
f_sub      = mkfont(BLACK, True, 11)
f_section  = mkfont(BLACK, True, 11, underline='single')
f_bold     = mkfont(BLACK, True)
f_normal   = mkfont(BLACK)
f_ital     = mkfont(BLACK, italic=True, size=9)
f_note     = mkfont('666666', italic=True, size=9)
f_blue     = mkfont(BLUE)
f_blue_b   = mkfont(BLUE, True)
f_green    = mkfont(GREEN)
f_green_b  = mkfont(GREEN, True)
f_black_b  = mkfont(BLACK, True)
f_result   = mkfont(BLACK, True, 12)
f_upside   = mkfont('0000CC', True, 11)
f_assum_h  = mkfont(BLACK, True, 12, underline='single')

# Number formats
pct   = '0.0%'
num   = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
dlr   = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
dec3  = '_(* #,##0.000_);_(* \\(#,##0.000\\);_(* "-"??_);_(@_)'
price = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
multx = '#,##0.0"x"'
bps_f = '#,##0'
ctr   = Alignment(horizontal='center')

thin = Side(style='thin')
bot_border  = Border(bottom=thin)
top_dbl     = Border(top=thin, bottom=Side(style='double'))

# Column widths
for c,w in {'A':2,'B':44,'C':5,'D':2,'E':2,'F':16,'G':16,'H':16,
            'I':16,'J':16,'K':16,'L':3,'M':30,'N':2,'O':2,'P':12,
            'Q':2,'R':38}.items():
    ws.column_dimensions[c].width = w

# ── Helper ───────────────────────────────────────────────────
def cell(addr, val, font=None, fmt=None, align=None, border=None):
    ws[addr] = val
    if font:   ws[addr].font = font
    if fmt:    ws[addr].number_format = fmt
    if align:  ws[addr].alignment = align
    if border: ws[addr].border = border

def model_ref(fy, row):
    """Return Model!{col}{row} reference string."""
    return f"=Model!{MC[fy]}{row}"

# ══════════════════════════════════════════════════════════════
# TITLE & HEADERS
# ══════════════════════════════════════════════════════════════
cell('B1', 'NVIDIA Corporation', f_title)
cell('B2', 'Discounted Cash Flow Valuation Analysis', f_sub)

cell('F5', 'Historical', mkfont(BLACK, True, 10), align=ctr)
cell('G5', 'Projected',  mkfont(BLACK, True, 10), align=ctr)
cell('M5', 'Assumptions', f_assum_h)
cell('R5', 'Notes', f_assum_h)

cell('B7', '($ in millions)', f_ital)
for fy in ALL:
    cell(f'{DC[fy]}7', fy, f_bold, align=ctr, border=bot_border)

# ══════════════════════════════════════════════════════════════
# ASSUMPTIONS PANEL (M6:R16) — same as before
# ══════════════════════════════════════════════════════════════
assumptions = [
    (6,  'Risk Free Rate:',             0.0425,  '0.00%', '10-Year US Treasury yield'),
    (7,  'Market Risk Premium:',        0.055,   '0.00%', 'Damodaran equity risk premium'),
    (8,  'Unlevered Beta:',             1.65,    '0.00',  'Semiconductor industry avg'),
    (9,  'Projected Equity to Value:',  0.998,   '0.0%',  'NVDA ~100% equity-financed'),
    (10, 'Levered Beta:',              '=(1+(1-P12)*(1-P9)/P9)*P8', '0.00', 'Hamada equation'),
    (11, 'Cost of Equity:',            '=P6+P7*P10', '0.00%', 'CAPM: Rf + B(Rm-Rf)'),
    (12, 'Tax Rate:',                   0.15,    '0%',    'Long-term effective tax rate'),
    (13, 'Cost of Debt:',              0.035,   '0.00%', 'Based on outstanding senior notes'),
    (14, 'WACC:',                      '=P11*P9+P13*(1-P12)*(1-P9)', '0.0%',
          'Rd*(1-T)*(D/V) + Re*(E/V)'),
    (15, 'Terminal Growth Rate:',       0.03,    '0.00%', 'Long-term nominal GDP growth'),
    (16, 'Implied TV/EBITDA:',         None,     multx,   'Terminal value / FY2030 EBITDA'),
]
for r, label, val, fmt, note in assumptions:
    cell(f'M{r}', label, f_bold)
    if val is not None:
        # Blue for hardcoded, black for formulas
        is_formula = isinstance(val, str) and val.startswith('=')
        cell(f'P{r}', val, f_blue if not is_formula else f_normal, fmt=fmt)
    cell(f'R{r}', note, f_note)

# ══════════════════════════════════════════════════════════════
# REVENUE BUILD (Rows 9–21)
# ══════════════════════════════════════════════════════════════
cell('B9', 'Revenue Build by Segment', f_section)

# Segments: (label, revenue_row, growth_row, dcf_row_rev, dcf_row_growth)
segments = [
    ('Datacenter Revenue',                22, 24, 10, 11),
    ('Gaming Revenue',                     7,  9, 12, 13),
    ('Automotive Revenue',                26, 28, 14, 15),
    ('Professional Visualization Revenue', 11, 13, 16, 17),
    ('OEM & IP Revenue',                  30, 32, 18, 19),
]

for label, mrev, mgrow, dr, dg in segments:
    cell(f'B{dr}', label, f_normal)
    cell(f'B{dg}', '  Y/Y Growth', f_normal)
    for fy in ALL:
        d = DC[fy]
        cell(f'{d}{dr}', model_ref(fy, mrev), f_green, fmt=num)
        cell(f'{d}{dg}', model_ref(fy, mgrow), f_green, fmt=pct)

# Row 20: Total Revenue
cell('B20', 'Total Revenue', f_green_b)
for fy in ALL:
    cell(f'{DC[fy]}20', model_ref(fy, 34), f_green_b, fmt=num, border=bot_border)

# Row 21: Y/Y Total Revenue Growth (computed locally)
cell('B21', 'Y/Y Total Revenue Growth', f_black_b)
for fy in PROJ:
    d = DC[fy]; p = DC[ALL[ALL.index(fy)-1]]
    cell(f'{d}21', f'={d}20/{p}20-1', f_normal, fmt=pct)

# ══════════════════════════════════════════════════════════════
# MARGIN / OPEX DRIVERS (Rows 23–32)
# ══════════════════════════════════════════════════════════════
cell('B23', 'Operating Margin Drivers', f_section)

# (label, model_row, dcf_row, fmt, font_override)
margin_lines = [
    ('Non-GAAP Gross Margin %',           59,  24, pct,   f_green),
    ('  Gross Margin Improvement (bps)',   62,  25, bps_f, f_green),
    ('Non-GAAP R&D Margin %',             76,  26, pct,   f_green),
    ('  R&D Margin Improvement (bps)',     77,  27, bps_f, f_green),
    ('Non-GAAP SG&A Margin %',            91,  28, pct,   f_green),
    ('  SG&A Margin Improvement (bps)',    92,  29, bps_f, f_green),
    ('Non-GAAP EBIT Margin %',           109,  30, pct,   f_green_b),
]

for label, mrow, drow, fmtv, fnt in margin_lines:
    cell(f'B{drow}', label, f_normal if fnt != f_green_b else f_bold)
    for fy in ALL:
        cell(f'{DC[fy]}{drow}', model_ref(fy, mrow), fnt, fmt=fmtv)

# Row 31: Non-GAAP EBIT
cell('B31', 'Non-GAAP Operating Income (EBIT)', f_green_b)
for fy in ALL:
    cell(f'{DC[fy]}31', model_ref(fy, 386), f_green_b, fmt=num, border=bot_border)

# ══════════════════════════════════════════════════════════════
# UFCF BUILD (Rows 33–40)
# ══════════════════════════════════════════════════════════════
cell('B33', 'Unlevered Free Cash Flow Build', f_section)

# Row 34: EBIT (local ref to row 31)
cell('B34', 'Non-GAAP EBIT', f_bold)
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}34', f'={d}31', f_normal, fmt=num)

# Row 35: Less Taxes
cell('B35', 'Less: Taxes on EBIT')
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}35', f'=-{d}34*$P$12', f_normal, fmt=num)

# Row 36: NOPAT
cell('B36', 'NOPAT', f_bold)
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}36', f'={d}34+{d}35', f_normal, fmt=num)

# Row 37: D&A
cell('B37', 'Plus: Depreciation & Amortization')
for fy in PROJ:
    cell(f'{DC[fy]}37', model_ref(fy, 933), f_green, fmt=num)

# Row 38: CapEx
cell('B38', 'Less: Capital Expenditures')
for fy in PROJ:
    cell(f'{DC[fy]}38', f"=-Model!{MC[fy]}934", f_green, fmt=num)

# Row 39: NWC (references row 72)
cell('B39', 'Less: Increase in Net Working Capital')
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}39', f'=-{d}72', f_normal, fmt=num)

# Row 40: UFCF
cell('B40', 'Unlevered Free Cash Flow', f_black_b, border=top_dbl)
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}40', f'=SUM({d}36:{d}39)', f_black_b, fmt=num, border=top_dbl)

# ══════════════════════════════════════════════════════════════
# DISCOUNTING (Rows 42–43)
# ══════════════════════════════════════════════════════════════
cell('B42', 'Times: Discount Factor')
for i, fy in enumerate(PROJ, 1):
    cell(f'{DC[fy]}42', f'=1/(1+$P$14)^{i}', f_normal, fmt=dec3)

cell('B43', 'Discounted Cash Flows')
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}43', f'={d}40*{d}42', f_normal, fmt=num)

# ══════════════════════════════════════════════════════════════
# VALUATION (Rows 45–49)
# ══════════════════════════════════════════════════════════════
cell('B45', 'Sum of Discounted Cash Flows', f_bold)
cell('F45', '=SUM(G43:K43)', f_normal, fmt=dlr)

cell('B46', 'Terminal Value:', f_bold)
cell('F46', '=(K40*(1+$P$15))/($P$14-$P$15)', f_normal, fmt=num)

cell('B47', 'Times: Discount Factor')
cell('F47', '=K42', f_normal, fmt=dec3)

cell('B48', 'PV of Terminal Value')
cell('F48', '=F46*F47', f_normal, fmt=dlr)

cell('C49', 'Enterprise Value', f_result)
cell('F49', '=F45+F48', f_normal, fmt=dlr, border=top_dbl)

# Update implied TV/EBITDA in assumptions panel
cell('P16', f"=F46/Model!{MC['FY2030']}392", f_green, fmt=multx)

# ══════════════════════════════════════════════════════════════
# EQUITY BRIDGE (Rows 51–61)
# ══════════════════════════════════════════════════════════════
cell('B51', 'Equity Bridge:', f_section)

cell('B52', 'Enterprise Value')
cell('F52', '=F49', f_normal, fmt=dlr)

cell('B53', 'Less: Total Debt')
cell('F53',
     f"=-(Model!{MC['FY2025']}992+IF(ISNUMBER(Model!{MC['FY2025']}988),Model!{MC['FY2025']}988,0))",
     f_green, fmt=dlr)

cell('B54', 'Plus: Cash & Cash Equivalents')
cell('F54', model_ref('FY2025', 952), f_green, fmt=dlr)

cell('B55', 'Plus: Marketable Securities')
cell('F55', model_ref('FY2025', 953), f_green, fmt=dlr)

cell('C56', 'Equity Value', f_result)
cell('F56', '=SUM(F52:F55)', f_normal, fmt=dlr, border=top_dbl)

cell('B57', 'Diluted Shares Outstanding (mm)')
cell('F57', model_ref('FY2025', 429), f_green, fmt=num)

cell('C58', 'Implied Share Price ($/share)', f_result)
cell('F58', '=F56/F57', f_normal, fmt=price, border=top_dbl)

cell('B60', 'Current Stock Price ($/share)', f_bold)
cell('F60', "='Front Page'!H20", f_green, fmt=price)

cell('B61', 'Upside / (Downside)', f_upside)
cell('F61', '=F58/F60-1', f_upside, fmt=pct)

# ══════════════════════════════════════════════════════════════
# NWC DETAIL (Rows 64–73)
# ══════════════════════════════════════════════════════════════
cell('B64', 'Net Working Capital Calculation:', f_section)

for fy in ALL:
    cell(f'{DC[fy]}65', fy, f_bold, align=ctr, border=bot_border)

nwc_items = [
    ('Accounts Receivable',                    954, 66, False),
    ('Inventories',                            958, 67, False),
    ('Less: Accounts Payable',                 985, 68, True),
    ('Less: Accrued & Other Current Liabilities', 987, 69, True),
]
for label, mrow, drow, negate in nwc_items:
    cell(f'B{drow}', label)
    for fy in ALL:
        sign = '-' if negate else ''
        cell(f'{DC[fy]}{drow}', f"={sign}Model!{MC[fy]}{mrow}", f_green, fmt=num)

cell('B70', 'Net Working Capital', f_bold, border=bot_border)
for fy in ALL:
    d = DC[fy]
    cell(f'{d}70', f'=SUM({d}66:{d}69)', f_normal, fmt=num, border=bot_border)

cell('B72', 'Change in Net Working Capital', f_bold)
for fy in PROJ:
    d = DC[fy]; p = DC[ALL[ALL.index(fy)-1]]
    cell(f'{d}72', f'={d}70-{p}70', f_normal, fmt=num)

# ══════════════════════════════════════════════════════════════
# MEMO (Rows 75–78)
# ══════════════════════════════════════════════════════════════
cell('B75', 'Memo:', f_section)

cell('B76', 'Stock-Based Compensation (excluded from UFCF)')
for fy in PROJ:
    cell(f'{DC[fy]}76', model_ref(fy, 408), f_green, fmt=num)

cell('B77', 'Non-GAAP EBITDA')
for fy in ALL:
    cell(f'{DC[fy]}77', model_ref(fy, 392), f_green, fmt=num)

cell('B78', 'UFCF Growth Rate')
for i, fy in enumerate(PROJ):
    if i > 0:
        d = DC[fy]; p = DC[PROJ[i-1]]
        cell(f'{d}78', f'={d}40/{p}40-1', f_normal, fmt=pct)

# ── Save ─────────────────────────────────────────────────────
wb.save('NVIDIA NVDA US.xlsx')
print('DCF tab rebuilt successfully!')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
