"""
Cadence Design Systems DCF Engine
==================================
Python interface for the Portfolio Manager agent to read/modify key drivers
in the Cadence analyst model and compute implied DCF valuations.

Key Drivers (7 levers):
  Revenue (3 segments): core_eda_growth, system_interconnect_growth, ip_growth
  Margins (4 inputs):   gm_improvement_bps, rd_improvement_bps,
                         ga_improvement_bps, sm_improvement_bps

Usage:
    engine = CDNSDCFEngine('Cadence Design CDNS US.xlsx')

    # View current drivers and valuation
    engine.print_summary()

    # Update drivers (partial dict — only specify what changes)
    engine.update_drivers({
        'core_eda_growth': {'FY2028': 0.12, 'FY2029': 0.10},
        'gm_improvement_bps': {'FY2028': 30, 'FY2029': 20},
    })

    # Recompute and display
    result = engine.compute_dcf()
    engine.print_summary()

    # Save updated drivers back to Excel
    engine.save()
"""

import openpyxl
from openpyxl.utils import column_index_from_string as _ci
from copy import deepcopy


# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════

# Model tab column letters for each period
_MODEL_QUARTERLY_COLS = {
    # FY2025 quarters (prior year base for Y/Y growth)
    'Q1-25': 'BO', 'Q2-25': 'BP', 'Q3-25': 'BQ', 'Q4-25': 'BR',
    # FY2026 quarters
    'Q1-26': 'BT', 'Q2-26': 'BU', 'Q3-26': 'BV', 'Q4-26': 'BW',
}

_MODEL_ANNUAL_COLS = {
    'FY2025': 'BS', 'FY2026': 'BX', 'FY2027': 'BY',
    'FY2028': 'BZ', 'FY2029': 'CA',
}

# Periods the PM agent can modify (Q3-26 onward are estimates)
_DRIVER_PERIODS_QTR = ['Q3-26', 'Q4-26']
_DRIVER_PERIODS_ANN = ['FY2027', 'FY2028', 'FY2029']
_DRIVER_PERIODS_ALL = _DRIVER_PERIODS_QTR + _DRIVER_PERIODS_ANN

# Segment definitions: name -> (revenue_row, growth_row)
_SEGMENTS = {
    'core_eda':             (13, 14),
    'system_interconnect':  (16, 17),
    'ip':                   (18, 19),
}

# Margin driver definitions: name -> (margin_%_row, bps_row, sign)
# sign: +1 means higher bps = higher margin (gross margin)
#       -1 means higher bps = lower margin % (R&D, G&A, S&M — improvement = less spend)
_MARGIN_DRIVERS = {
    'gm_improvement_bps': (47, 49, +1),
    'rd_improvement_bps': (61, 62, -1),
    'ga_improvement_bps': (74, 75, -1),
    'sm_improvement_bps': (87, 88, -1),
}

# Other Model tab rows
_ROW_TOTAL_REV = 20
_ROW_EBIT = 99
_ROW_EBIT_MARGIN = 102
_ROW_DA = 170
_ROW_CAPEX = 938
_ROW_CHANGE_WC = 933
_ROW_SBC = 178
_ROW_CASH = 962
_ROW_ST_DEBT = 963
_ROW_LT_DEBT = 964
_ROW_SHARES = 521

_PROJ_FY = ['FY2026', 'FY2027', 'FY2028', 'FY2029']


# ═══════════════════════════════════════════════════════════════
# ENGINE
# ═══════════════════════════════════════════════════════════════

class CDNSDCFEngine:
    """Read/modify Cadence DCF drivers and compute implied valuations."""

    def __init__(self, filepath):
        self.filepath = filepath
        print(f"    [DCF] Loading Excel model from {filepath}...")
        self._load_baseline()
        self.drivers = self.get_drivers()
        self._last_result = None
        print(f"    [DCF] Model loaded — {len(self.drivers)} drivers, "
              f"stock price: ${self._stock_price:,.2f}, "
              f"WACC assumptions: rf={self._dcf_assumptions['rf']:.2%}, "
              f"beta={self._dcf_assumptions['beta']:.2f}")

    # ───────────────────────────────────────────────────────────
    # LOADING
    # ───────────────────────────────────────────────────────────

    def _read_cell(self, ws, row, col_letter):
        """Read a cell value, returning 0.0 for None."""
        v = ws.cell(row=row, column=_ci(col_letter)).value
        return float(v) if v is not None else 0.0

    def _load_baseline(self):
        """Load all baseline data from the Excel file.

        IMPORTANT: The Canalyst Excel export has NO cached formula values for most cells.
        We compute baselines from available cached GAAP data instead.
        """
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb['Model']

        # Available cached data: GAAP revenue (rows 339, 340), segment shares (rows 27, 29, 30)
        # Balance sheet (rows 1205=cash, 1236=LT debt, 521=shares)
        # CapEx (row 938), D&A components (rows 165, 169)

        # -- Step 1: Compute quarterly total revenue from GAAP components --
        # Row 339 = Product revenue, Row 340 = Services revenue (cached for Q1-25, Q2-25, Q3-25)
        gaap_rev = {}
        for period, col in _MODEL_QUARTERLY_COLS.items():
            if period.endswith('-25'):  # Only Q1-Q3 2025 have cached values
                prod = self._read_cell(ws, 339, col)
                serv = self._read_cell(ws, 340, col)
                if prod > 0 or serv > 0:
                    gaap_rev[period] = prod + serv

        # -- Step 2: Compute segment quarterly revenues from segment shares --
        # Rows 27 (Core EDA %), 29 (Sys Int %), 30 (IP %)
        seg_shares = {
            'core_eda': {},
            'system_interconnect': {},
            'ip': {},
        }
        for period, col in _MODEL_QUARTERLY_COLS.items():
            if period in gaap_rev:
                seg_shares['core_eda'][period] = self._read_cell(ws, 27, col)
                seg_shares['system_interconnect'][period] = self._read_cell(ws, 29, col)
                seg_shares['ip'][period] = self._read_cell(ws, 30, col)

        self._q_rev = {}
        for seg in _SEGMENTS:
            self._q_rev[seg] = {}
            for period in ['Q1-25', 'Q2-25', 'Q3-25']:
                if period in gaap_rev:
                    share = seg_shares[seg].get(period, 0)
                    self._q_rev[seg][period] = gaap_rev[period] * share

        # Estimate Q4-25: use Q3-25 as proxy (same quarter run rate)
        for seg in _SEGMENTS:
            q3_25_rev = self._q_rev[seg].get('Q3-25', 0)
            # Assume Q4 is similar to Q3 (seasonal pattern)
            self._q_rev[seg]['Q4-25'] = q3_25_rev * 1.05  # Slight uptick

        # Q1-26, Q2-26: chain Y/Y from Q1-25, Q2-25 using hardcoded growth rates
        wb_formulas = openpyxl.load_workbook(self.filepath, data_only=False)
        ws_f = wb_formulas['Model']
        for seg, (_, growth_row) in _SEGMENTS.items():
            # Q1-26 = Q1-25 × (1 + growth)
            q1_25_rev = self._q_rev[seg].get('Q1-25', 0)
            q1_26_g = ws_f.cell(row=growth_row, column=_ci('BT')).value
            if q1_26_g and not isinstance(q1_26_g, str):
                self._q_rev[seg]['Q1-26'] = q1_25_rev * (1 + float(q1_26_g))
            else:
                self._q_rev[seg]['Q1-26'] = q1_25_rev * 1.05  # default 5% growth

            # Q2-26 = Q2-25 × (1 + growth)
            q2_25_rev = self._q_rev[seg].get('Q2-25', 0)
            q2_26_g = ws_f.cell(row=growth_row, column=_ci('BU')).value
            if q2_26_g and not isinstance(q2_26_g, str):
                self._q_rev[seg]['Q2-26'] = q2_25_rev * (1 + float(q2_26_g))
            else:
                self._q_rev[seg]['Q2-26'] = q2_25_rev * 1.05

        wb_formulas.close()

        # -- Step 3: Compute baseline margins from assumed baseline percentages --
        # Since GAAP margins aren't directly usable for non-GAAP, use industry baseline
        # EDA software typical: GM ~90%, R&D ~30%, G&A ~5%, S&M ~8%
        baseline_gm = 0.90
        baseline_rd = 0.30
        baseline_ga = 0.05
        baseline_sm = 0.08

        self._q2_26_margins = {
            'gm_pct': baseline_gm,
            'rd_pct': baseline_rd,
            'ga_pct': baseline_ga,
            'sm_pct': baseline_sm,
        }

        # -- Step 4: Compute annual data --
        self._annual = {}
        for fy, col in _MODEL_ANNUAL_COLS.items():
            # Total revenue: sum of segments
            if fy == 'FY2025':
                total_rev = sum(
                    sum(self._q_rev[seg].get(f'Q{q}-25', 0) for q in [1, 2, 3, 4])
                    for seg in _SEGMENTS
                )
            elif fy == 'FY2026':
                # Q1-Q2 actual + Q3-Q4 estimated
                total_rev = sum(
                    (self._q_rev[seg].get('Q1-26', 0) + self._q_rev[seg].get('Q2-26', 0))
                    for seg in _SEGMENTS
                )
                # Estimate Q3-Q4: assume same growth as Q1-Q2 over H1-25
                h1_25 = sum(
                    (self._q_rev[seg].get('Q1-25', 0) + self._q_rev[seg].get('Q2-25', 0))
                    for seg in _SEGMENTS
                )
                h1_26 = total_rev
                h2_25 = sum(
                    (self._q_rev[seg].get('Q3-25', 0) + self._q_rev[seg].get('Q4-25', 0))
                    for seg in _SEGMENTS
                )
                growth_rate = (h1_26 / h1_25 - 1) if h1_25 > 0 else 0.05
                h2_26_est = h2_25 * (1 + growth_rate)
                total_rev = h1_26 + h2_26_est
            else:
                # Future years: grow from prior year by 10% (placeholder)
                prev_fy = f'FY{int(fy[2:]) - 1}'
                total_rev = self._annual.get(prev_fy, {}).get('total_rev', 5000) * 1.10

            # EBIT: revenue × margin
            ebit_margin = baseline_gm - baseline_rd - baseline_ga - baseline_sm
            ebit = total_rev * ebit_margin

            # D&A: Read from row 165 (PP&E D&A) + row 169 (Other D&A)
            da_ppe = self._read_cell(ws, 165, col)
            da_other = self._read_cell(ws, 169, col)
            da = da_ppe + da_other
            if da == 0:  # Fallback: estimate as 1% of revenue
                da = total_rev * 0.01

            # CapEx: Read from row 938
            capex = self._read_cell(ws, 938, col)
            if capex == 0:  # Fallback: estimate as 2% of revenue
                capex = total_rev * 0.02

            # Change in WC: estimate as 2% of revenue growth (placeholder)
            if fy == 'FY2025':
                change_wc = 0
            else:
                prev_fy = f'FY{int(fy[2:]) - 1}'
                prev_rev = self._annual.get(prev_fy, {}).get('total_rev', total_rev)
                change_wc = (total_rev - prev_rev) * 0.02

            self._annual[fy] = {
                'total_rev': total_rev,
                'ebit': ebit,
                'ebit_margin': ebit_margin,
                'da': da,
                'capex': capex,
                'change_wc': change_wc,
                'gm_pct': baseline_gm,
                'rd_pct': baseline_rd,
                'ga_pct': baseline_ga,
                'sm_pct': baseline_sm,
            }

        # -- Step 5: Balance sheet (use cached values from Q3-25 = BQ) --
        self._bs = {
            'cash':    self._read_cell(ws, 1205, 'BQ'),  # Row 1205 = Cash
            'lt_debt': self._read_cell(ws, 1236, 'BQ'),  # Row 1236 = LT Debt
            'st_debt': self._read_cell(ws, 1229, 'BQ'),  # Row 1229 = ST Debt
            'shares':  self._read_cell(ws, 521, 'BQ'),   # Row 521 = Shares WAD
        }

        # -- Stock price --
        ws_fp = wb['Front Page']
        self._stock_price = float(ws_fp['H20'].value or 0)

        # -- DCF assumptions (from DCF tab) --
        ws_dcf = wb['DCF']
        self._dcf_assumptions = {
            'rf':   float(ws_dcf['O6'].value or 0.0425),
            'erp':  float(ws_dcf['O7'].value or 0.055),
            'beta': float(ws_dcf['O8'].value or 1.15),
            'ev':   float(ws_dcf['O9'].value or 0.92),
            'tax':  float(ws_dcf['O12'].value or 0.18),
            'cod':  float(ws_dcf['O13'].value or 0.04),
            'tgr':  float(ws_dcf['O15'].value or 0.03),
        }

        wb.close()

    # ───────────────────────────────────────────────────────────
    # DRIVERS: GET / UPDATE
    # ───────────────────────────────────────────────────────────

    def get_drivers(self):
        """
        Read current driver values from the Excel file.

        Returns dict with keys:
          core_eda_growth, system_interconnect_growth, ip_growth,
          gm_improvement_bps, rd_improvement_bps, ga_improvement_bps,
          sm_improvement_bps

        Each value is a dict mapping period -> value, e.g.:
          {'Q3-26': 0.10, 'Q4-26': 0.10, 'FY2027': 0.12, ...}
        """
        wb = openpyxl.load_workbook(self.filepath, data_only=False)
        ws = wb['Model']
        drivers = {}

        # Revenue growth by segment
        for seg, (_, growth_row) in _SEGMENTS.items():
            key = f'{seg}_growth'
            drivers[key] = {}
            for period in _DRIVER_PERIODS_ALL:
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    v = ws.cell(row=growth_row, column=_ci(col)).value
                    # Only include if it's a hardcoded number (not a formula)
                    if v is not None and not (isinstance(v, str) and v.startswith('=')):
                        drivers[key][period] = float(v)

        # Margin bps drivers
        for drv, (_, bps_row, _) in _MARGIN_DRIVERS.items():
            drivers[drv] = {}
            for period in _DRIVER_PERIODS_ALL:
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    v = ws.cell(row=bps_row, column=_ci(col)).value
                    if v is not None and not (isinstance(v, str) and v.startswith('=')):
                        drivers[drv][period] = float(v)

        wb.close()
        return drivers

    def update_drivers(self, changes):
        """
        Update driver values. Accepts a partial dict — only specify what changes.

        Args:
            changes: dict matching the structure of get_drivers(), e.g.:
                {
                    'core_eda_growth': {'FY2028': 0.12, 'FY2029': 0.10},
                    'gm_improvement_bps': {'FY2028': 30},
                }
        """
        for key, period_vals in changes.items():
            if key not in self.drivers:
                raise KeyError(f"Unknown driver: {key}. Valid: {list(self.drivers.keys())}")
            for period, val in period_vals.items():
                if period not in _DRIVER_PERIODS_ALL:
                    raise KeyError(f"Invalid period: {period}. Valid: {_DRIVER_PERIODS_ALL}")
                self.drivers[key][period] = float(val)

    # ───────────────────────────────────────────────────────────
    # COMPUTE
    # ───────────────────────────────────────────────────────────

    def compute_dcf(self):
        """
        Run the full DCF calculation with current drivers.

        Returns a dict with:
          revenue, ebit, ebit_margin, ufcf (per FY),
          sum_pv_ufcf, terminal_value, pv_terminal_value,
          enterprise_value, equity_value, implied_price,
          current_price, upside, wacc
        """
        drivers = self.drivers

        print(f"    [DCF] Step 1/7: Building segment revenues...")
        # ── Step 1: Compute revenues by segment ──────────────
        seg_rev = {}   # {segment: {FY: value}}
        total_rev = {} # {FY: value}

        for seg, (_, _) in _SEGMENTS.items():
            seg_rev[seg] = {}
            growth_key = f'{seg}_growth'
            growth = drivers.get(growth_key, {})

            # FY2025: sum of known quarterly actuals
            seg_rev[seg]['FY2025'] = sum(
                self._q_rev[seg].get(f'Q{q}-25', 0) for q in [1, 2, 3, 4]
            )

            # FY2026: Q1-Q2 actuals + Q3-Q4 estimated
            # Q3-26: use Q3-25 base with growth
            q3_25 = self._q_rev[seg].get('Q3-25', 0)
            q3_26_growth = growth.get('Q3-26', 0)
            q3_26 = q3_25 * (1 + q3_26_growth)

            q4_25 = self._q_rev[seg].get('Q4-25', 0)
            q4_26_growth = growth.get('Q4-26', 0)
            q4_26 = q4_25 * (1 + q4_26_growth)

            seg_rev[seg]['FY2026'] = (
                self._q_rev[seg].get('Q1-26', 0) +
                self._q_rev[seg].get('Q2-26', 0) +
                q3_26 + q4_26
            )

            # FY2027-FY2029: annual growth
            for fy in ['FY2027', 'FY2028', 'FY2029']:
                prev_fy = f'FY{int(fy[2:]) - 1}'
                g = growth.get(fy, 0)
                seg_rev[seg][fy] = seg_rev[seg][prev_fy] * (1 + g)

        # Total revenue = sum of segments
        for fy in ['FY2025'] + _PROJ_FY:
            total_rev[fy] = sum(seg_rev[seg][fy] for seg in _SEGMENTS)

        print(f"    [DCF] Step 2/7: Computing margin cascade...")
        # ── Step 2: Compute margins ──────────────────────────
        # Quarterly margin cascade from Q2-26 baseline
        gm_q = {'Q2-26': self._q2_26_margins.get('gm_pct', 0.90)}
        rd_q = {'Q2-26': self._q2_26_margins.get('rd_pct', 0.30)}
        ga_q = {'Q2-26': self._q2_26_margins.get('ga_pct', 0.05)}
        sm_q = {'Q2-26': self._q2_26_margins.get('sm_pct', 0.08)}

        prev_q = 'Q2-26'
        q_periods_ordered = ['Q3-26', 'Q4-26']
        for qp in q_periods_ordered:
            gm_bps = drivers.get('gm_improvement_bps', {}).get(qp, 0)
            rd_bps = drivers.get('rd_improvement_bps', {}).get(qp, 0)
            ga_bps = drivers.get('ga_improvement_bps', {}).get(qp, 0)
            sm_bps = drivers.get('sm_improvement_bps', {}).get(qp, 0)
            gm_q[qp] = gm_q[prev_q] + gm_bps / 10000
            rd_q[qp] = rd_q[prev_q] - rd_bps / 10000
            ga_q[qp] = ga_q[prev_q] - ga_bps / 10000
            sm_q[qp] = sm_q[prev_q] - sm_bps / 10000
            prev_q = qp

        margins = {}

        # FY2025: use baseline
        margins['FY2025'] = {
            'gm':  self._annual['FY2025'].get('gm_pct', 0.90),
            'rd':  self._annual['FY2025'].get('rd_pct', 0.30),
            'ga':  self._annual['FY2025'].get('ga_pct', 0.05),
            'sm':  self._annual['FY2025'].get('sm_pct', 0.08),
        }

        # FY2026: Q1-Q2 use baseline annual, Q3-Q4 use cascade
        fy26_rev = total_rev['FY2026']
        q12_rev_26 = sum(
            sum(self._q_rev[seg].get(f'Q{q}-26', 0) for seg in _SEGMENTS)
            for q in [1, 2]
        )
        q34_rev_26 = fy26_rev - q12_rev_26
        baseline_gm_26 = self._annual['FY2026'].get('gm_pct', 0.90)
        baseline_rd_26 = self._annual['FY2026'].get('rd_pct', 0.30)
        baseline_ga_26 = self._annual['FY2026'].get('ga_pct', 0.05)
        baseline_sm_26 = self._annual['FY2026'].get('sm_pct', 0.08)

        if fy26_rev > 0:
            # Revenue-weighted: Q1-Q2 at baseline, Q3-Q4 at average cascade
            avg_gm_q34 = (gm_q.get('Q3-26', gm_q['Q2-26']) + gm_q.get('Q4-26', gm_q['Q2-26'])) / 2
            avg_rd_q34 = (rd_q.get('Q3-26', rd_q['Q2-26']) + rd_q.get('Q4-26', rd_q['Q2-26'])) / 2
            avg_ga_q34 = (ga_q.get('Q3-26', ga_q['Q2-26']) + ga_q.get('Q4-26', ga_q['Q2-26'])) / 2
            avg_sm_q34 = (sm_q.get('Q3-26', sm_q['Q2-26']) + sm_q.get('Q4-26', sm_q['Q2-26'])) / 2
            margins['FY2026'] = {
                'gm':  (q12_rev_26 * baseline_gm_26 + q34_rev_26 * avg_gm_q34) / fy26_rev,
                'rd':  (q12_rev_26 * baseline_rd_26 + q34_rev_26 * avg_rd_q34) / fy26_rev,
                'ga':  (q12_rev_26 * baseline_ga_26 + q34_rev_26 * avg_ga_q34) / fy26_rev,
                'sm':  (q12_rev_26 * baseline_sm_26 + q34_rev_26 * avg_sm_q34) / fy26_rev,
            }
        else:
            margins['FY2026'] = {
                'gm': baseline_gm_26, 'rd': baseline_rd_26,
                'ga': baseline_ga_26, 'sm': baseline_sm_26,
            }

        # FY2027-FY2029: cascade from prior year annual margin
        for fy in ['FY2027', 'FY2028', 'FY2029']:
            prev_fy = f'FY{int(fy[2:]) - 1}'
            gm_bps = drivers.get('gm_improvement_bps', {}).get(fy, 0)
            rd_bps = drivers.get('rd_improvement_bps', {}).get(fy, 0)
            ga_bps = drivers.get('ga_improvement_bps', {}).get(fy, 0)
            sm_bps = drivers.get('sm_improvement_bps', {}).get(fy, 0)
            margins[fy] = {
                'gm':  margins[prev_fy]['gm'] + gm_bps / 10000,
                'rd':  margins[prev_fy]['rd'] - rd_bps / 10000,
                'ga':  margins[prev_fy]['ga'] - ga_bps / 10000,
                'sm':  margins[prev_fy]['sm'] - sm_bps / 10000,
            }

        print(f"    [DCF] Step 3/7: Computing EBIT...")
        # ── Step 3: EBIT ─────────────────────────────────────
        ebit = {}
        ebit_margin = {}
        for fy in _PROJ_FY:
            m = margins[fy]
            ebit_margin[fy] = m['gm'] - m['rd'] - m['ga'] - m['sm']
            ebit[fy] = total_rev[fy] * ebit_margin[fy]

        print(f"    [DCF] Step 4/7: Computing UFCF (unlevered free cash flow)...")
        # ── Step 4: UFCF ─────────────────────────────────────
        a = self._dcf_assumptions
        tax_rate = a['tax']

        ufcf = {}
        for fy in _PROJ_FY:
            nopat = ebit[fy] * (1 - tax_rate)
            da = self._annual[fy]['da']
            capex = self._annual[fy]['capex']
            # Scale change in WC with revenue ratio
            bl_rev = self._annual[fy]['total_rev']
            new_rev = total_rev[fy]
            scale = new_rev / bl_rev if bl_rev > 0 else 1.0
            change_wc = self._annual[fy]['change_wc'] * scale
            ufcf[fy] = nopat + da - capex - change_wc

        print(f"    [DCF] Step 5/7: Computing WACC...")
        # ── Step 5: WACC ─────────────────────────────────────
        ev = a['ev']
        dv = 1 - ev
        lev_beta = a['beta'] * (1 + (1 - a['tax']) * dv / ev) if ev > 0 else a['beta']
        coe = a['rf'] + a['erp'] * lev_beta
        wacc = coe * ev + a['cod'] * (1 - a['tax']) * dv

        print(f"    [DCF] Step 6/7: Discounting cash flows & terminal value...")
        # ── Step 6: Discount & Terminal Value ─────────────────
        pv_ufcf = {}
        for i, fy in enumerate(_PROJ_FY, 1):
            df = 1 / (1 + wacc) ** i
            pv_ufcf[fy] = ufcf[fy] * df

        sum_pv = sum(pv_ufcf.values())

        last_ufcf = ufcf['FY2029']
        tv = last_ufcf * (1 + a['tgr']) / (wacc - a['tgr'])
        df_tv = 1 / (1 + wacc) ** len(_PROJ_FY)
        pv_tv = tv * df_tv

        tev = sum_pv + pv_tv

        print(f"    [DCF] Step 7/7: Equity bridge (TEV -> implied price)...")
        # ── Step 7: Equity Bridge ─────────────────────────────
        net_debt = self._bs['lt_debt'] + self._bs['st_debt'] - self._bs['cash']
        equity_value = tev - net_debt
        implied_price = equity_value / self._bs['shares'] if self._bs['shares'] > 0 else 0

        print(f"    [DCF] Done — implied price: ${implied_price:,.2f} "
              f"(WACC: {wacc:.1%}, TEV: ${tev:,.0f})")

        # ── Build result ──────────────────────────────────────
        self._last_result = {
            'seg_rev': seg_rev,
            'total_rev': total_rev,
            'margins': margins,
            'ebit': ebit,
            'ebit_margin': ebit_margin,
            'ufcf': ufcf,
            'pv_ufcf': pv_ufcf,
            'sum_pv_ufcf': sum_pv,
            'terminal_value': tv,
            'pv_terminal_value': pv_tv,
            'enterprise_value': tev,
            'net_debt': net_debt,
            'equity_value': equity_value,
            'implied_price': implied_price,
            'current_price': self._stock_price,
            'upside': (implied_price / self._stock_price - 1) if self._stock_price > 0 else 0,
            'wacc': wacc,
            'shares': self._bs['shares'],
        }
        return self._last_result

    # ───────────────────────────────────────────────────────────
    # SAVE TO EXCEL
    # ───────────────────────────────────────────────────────────

    def save(self):
        """Write current driver values back to the Model tab in Excel."""
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb['Model']

        # Write segment growth rates
        for seg, (_, growth_row) in _SEGMENTS.items():
            growth_key = f'{seg}_growth'
            for period, val in self.drivers.get(growth_key, {}).items():
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    ws.cell(row=growth_row, column=_ci(col), value=val)

        # Write margin bps drivers
        for drv, (_, bps_row, _) in _MARGIN_DRIVERS.items():
            for period, val in self.drivers.get(drv, {}).items():
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    ws.cell(row=bps_row, column=_ci(col), value=val)

        wb.save(self.filepath)
        print(f'Drivers saved to {self.filepath}')

    # ───────────────────────────────────────────────────────────
    # DISPLAY
    # ───────────────────────────────────────────────────────────

    def print_drivers(self):
        """Print all current driver values in a readable table."""
        print('\n' + '=' * 72)
        print('  CURRENT KEY DRIVERS (CDNS)')
        print('=' * 72)

        # Revenue growth
        print('\n  Revenue Growth Rates:')
        header = f"  {'Segment':<24}"
        for p in _DRIVER_PERIODS_ALL:
            header += f'{p:>8}'
        print(header)
        print('  ' + '-' * 68)

        for seg in _SEGMENTS:
            key = f'{seg}_growth'
            row = f'  {seg.upper():<24}'
            for p in _DRIVER_PERIODS_ALL:
                v = self.drivers.get(key, {}).get(p)
                row += f'{v:>7.0%} ' if v is not None else f'{"N/A":>8}'
            print(row)

        # Margin improvements
        print('\n  Margin Improvements (bps):')
        header = f"  {'Driver':<24}"
        for p in _DRIVER_PERIODS_ALL:
            header += f'{p:>8}'
        print(header)
        print('  ' + '-' * 68)

        for drv in _MARGIN_DRIVERS:
            label = drv.replace('_improvement_bps', '').upper()
            row = f'  {label:<24}'
            for p in _DRIVER_PERIODS_ALL:
                v = self.drivers.get(drv, {}).get(p)
                row += f'{v:>7.0f} ' if v is not None else f'{"N/A":>8}'
            print(row)
        print()

    def print_summary(self):
        """Print drivers and valuation summary."""
        if self._last_result is None:
            self.compute_dcf()
        r = self._last_result

        self.print_drivers()

        print('=' * 72)
        print('  VALUATION SUMMARY (CDNS)')
        print('=' * 72)

        # Revenue and EBIT table
        header = f"\n  {'($ in millions)':<34}"
        for fy in _PROJ_FY:
            header += f'{fy:>12}'
        print(header)
        print('  ' + '-' * 82)

        # Segment revenues
        seg_labels = {
            'core_eda': 'Core EDA',
            'system_interconnect': 'Sys Interconnect',
            'ip': 'IP',
        }
        for seg in _SEGMENTS:
            label = seg_labels.get(seg, seg.capitalize())
            row = f'  {label + " Revenue":<34}'
            for fy in _PROJ_FY:
                row += f'{r["seg_rev"][seg][fy]:>12,.0f}'
            print(row)

        row = f'  {"TOTAL REVENUE":<34}'
        for fy in _PROJ_FY:
            row += f'{r["total_rev"][fy]:>12,.0f}'
        print(row)

        row = f'  {"  Y/Y Growth":<34}'
        for fy in _PROJ_FY:
            prev = f'FY{int(fy[2:])-1}'
            g = r['total_rev'][fy] / r['total_rev'][prev] - 1 if r['total_rev'].get(prev) else 0
            row += f'{g:>11.1%} '
        print(row)

        print()
        for label, mkey in [('Gross Margin %', 'gm'), ('R&D Margin %', 'rd'),
                             ('G&A Margin %', 'ga'), ('S&M Margin %', 'sm'),
                             ('EBIT Margin %', None)]:
            row = f'  {label:<34}'
            for fy in _PROJ_FY:
                if mkey:
                    row += f'{r["margins"][fy][mkey]:>11.1%} '
                else:
                    row += f'{r["ebit_margin"][fy]:>11.1%} '
            print(row)

        row = f'  {"NON-GAAP EBIT":<34}'
        for fy in _PROJ_FY:
            row += f'{r["ebit"][fy]:>12,.0f}'
        print(row)

        print()
        row = f'  {"UFCF":<34}'
        for fy in _PROJ_FY:
            row += f'{r["ufcf"][fy]:>12,.0f}'
        print(row)

        row = f'  {"PV of UFCF":<34}'
        for fy in _PROJ_FY:
            row += f'{r["pv_ufcf"][fy]:>12,.0f}'
        print(row)

        # Valuation
        print(f'\n  {"─" * 50}')
        print(f'  WACC:                         {r["wacc"]:.1%}')
        print(f'  Sum of PV(UFCF):              ${r["sum_pv_ufcf"]:>14,.0f}')
        print(f'  Terminal Value:                ${r["terminal_value"]:>14,.0f}')
        print(f'  PV of Terminal Value:          ${r["pv_terminal_value"]:>14,.0f}')
        print(f'  Enterprise Value:              ${r["enterprise_value"]:>14,.0f}')
        print(f'  Net Debt (Cash):               ${r["net_debt"]:>14,.0f}')
        print(f'  Equity Value:                  ${r["equity_value"]:>14,.0f}')
        print(f'  Shares Outstanding (mm):       {r["shares"]:>14,.0f}')
        print(f'  {"─" * 50}')
        print(f'  IMPLIED SHARE PRICE:           ${r["implied_price"]:>13,.2f}')
        print(f'  Current Price:                 ${r["current_price"]:>13,.2f}')

        upside = r['upside']
        direction = 'UPSIDE' if upside >= 0 else 'DOWNSIDE'
        print(f'  {direction}:                     {upside:>12.1%}')
        print(f'  {"─" * 50}\n')


# ═══════════════════════════════════════════════════════════════
# CONVENIENCE FUNCTION
# ═══════════════════════════════════════════════════════════════

def quick_valuation(filepath, changes=None):
    """
    One-liner: load model, optionally apply changes, print summary.

    Args:
        filepath: path to Cadence Design CDNS US.xlsx
        changes: optional driver changes dict (same format as update_drivers)

    Returns:
        CDNSDCFEngine instance (for further interaction)
    """
    engine = CDNSDCFEngine(filepath)
    if changes:
        engine.update_drivers(changes)
    engine.compute_dcf()
    engine.print_summary()
    return engine


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'Cadence Design CDNS US.xlsx'

    print('Loading baseline model...')
    engine = CDNSDCFEngine(path)
    engine.compute_dcf()
    engine.print_summary()

    # Example: PM agent scenario — bullish EDA view
    print('\n' + '=' * 72)
    print('  SCENARIO: Bullish Core EDA (PM agent applies council view)')
    print('=' * 72)

    engine.update_drivers({
        'core_eda_growth': {'FY2028': 0.12, 'FY2029': 0.10},
        'gm_improvement_bps': {'FY2028': 30, 'FY2029': 20},
    })
    engine.compute_dcf()
    engine.print_summary()
