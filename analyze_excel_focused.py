#!/usr/bin/env python3
"""
Focused analysis of Excel financial models to identify key DCF cells.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import json

def find_rows_with_keywords(ws, keywords, start_row=1, end_row=None):
    """Find rows containing specific keywords."""
    if end_row is None:
        end_row = ws.max_row

    results = {}
    for row_idx in range(start_row, min(end_row + 1, ws.max_row + 1)):
        for col_idx in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                cell_text = str(cell.value).lower()
                for keyword in keywords:
                    if keyword.lower() in cell_text:
                        col_letter = get_column_letter(col_idx)
                        key = f"Row {row_idx}"
                        if key not in results:
                            results[key] = []
                        results[key].append({
                            'cell': f"{col_letter}{row_idx}",
                            'value': str(cell.value)[:60],
                            'keyword': keyword
                        })
    return results

def analyze_company_model(filepath, company_name):
    """Analyze a company's financial model."""
    print(f"\n{'='*100}")
    print(f"COMPANY: {company_name}")
    print(f"File: {os.path.basename(filepath)}")
    print(f"{'='*100}\n")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        print(f"Sheets: {wb.sheetnames}\n")

        if 'Model' not in wb.sheetnames:
            print("ERROR: No 'Model' sheet found")
            return

        ws = wb['Model']
        print(f"Model sheet dimensions: {ws.dimensions}")
        print(f"Max rows: {ws.max_row}, Max columns: {ws.max_column}\n")

        # Search for key metrics
        keywords = [
            'revenue', 'product', 'gross', 'margin', 'operating', 'ebit',
            'ebitda', 'net income', 'earnings', 'depreciation', 'amortization',
            'capex', 'capital', 'cash flow', 'working capital', 'nwc',
            'accounts', 'inventory', 'payable', 'debt', 'cash',
            'share', 'diluted', 'tax', 'provision', 'income tax',
            'r&d', 'research', 'sga', 'selling', 'administrative',
            'segment', 'gaming', 'datacenter', 'automotive'
        ]

        print("=" * 100)
        print("KEY METRICS FOUND (Rows 1-300)")
        print("=" * 100)
        results = find_rows_with_keywords(ws, keywords, 1, 300)

        # Display results organized by row
        for row_key in sorted(results.keys(), key=lambda x: int(x.split()[1])):
            print(f"\n{row_key}:")
            for item in results[row_key]:
                print(f"  {item['cell']}: {item['value']}")

        # Check column structure in rows 1-10
        print("\n" + "=" * 100)
        print("COLUMN HEADER STRUCTURE (Rows 1-10)")
        print("=" * 100)

        for row_idx in range(1, 11):
            row_str = f"Row {row_idx}: "
            cols_with_data = []
            for col_idx in range(1, min(ws.max_column + 1, 50)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value:
                    col_letter = get_column_letter(col_idx)
                    cols_with_data.append(f"{col_letter}={str(cell.value)[:30]}")
            if cols_with_data:
                print(row_str + " | ".join(cols_with_data))

        wb.close()

    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

# Analyze companies
companies = {
    'TSMC': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/Taiwan Semiconductor Manufacturing Company TSM US.xlsx',
    'CoreWeave': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/CoreWeave CRWV US.xlsx',
    'ASML': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/ASML Holding ASML NA.xlsx',
}

for company, filepath in companies.items():
    if os.path.exists(filepath):
        analyze_company_model(filepath, company)
