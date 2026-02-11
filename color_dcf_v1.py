"""
color_dcf_v1.py
===============
Color coding script for the INITIAL (v1) DCF tab layout.
Kept for reference. The current build_dcf_tab.py already applies colors inline.

Convention:
  Blue  (0,0,255) = Hardcoded assumption inputs
  Green (0,127,0) = Cells pulling directly from another tab
  Black (0,0,0)   = Local DCF formulas

NOTE: This script was designed for the v1 layout (rows 9-57).
      The current DCF layout (rows 9-78) has colors baked into build_dcf_tab.py.
"""
import openpyxl
from openpyxl.styles import Font, Border, Side
from copy import copy

wb = openpyxl.load_workbook('NVIDIA NVDA US.xlsx')
ws = wb['DCF']

BLUE = '0000FF'
GREEN = '007F00'
BLACK = '000000'

def recolor(addr, color):
    """Change only the font color, preserving all other font attributes."""
    c = ws[addr]
    old = c.font
    c.font = Font(
        name=old.name, size=old.size, bold=old.bold, italic=old.italic,
        underline=old.underline, strike=old.strike,
        color=color
    )

# ── BLUE: Hardcoded assumption inputs ────────────────────────
blue_cells = ['P6','P7','P8','P9','P12','P13','P15']
for addr in blue_cells:
    recolor(addr, BLUE)

# ── GREEN: Cells that pull directly from another tab ─────────
# Revenue (Model tab)
for col in ['F','G','H','I','J','K']:
    recolor(f'{col}9', GREEN)    # Revenue
    recolor(f'{col}12', GREEN)   # EBIT

# D&A, CapEx (projected only, from Model)
for col in ['G','H','I','J','K']:
    recolor(f'{col}16', GREEN)   # D&A
    recolor(f'{col}17', GREEN)   # CapEx

# NWC items (all years, from Model)
for col in ['F','G','H','I','J','K']:
    recolor(f'{col}45', GREEN)   # AR
    recolor(f'{col}46', GREEN)   # Inventory
    recolor(f'{col}47', GREEN)   # AP
    recolor(f'{col}48', GREEN)   # Accrued

# Equity bridge items (from Model, FY2025 BS)
for addr in ['F32','F33','F34','F36']:
    recolor(addr, GREEN)

# Current stock price (from Front Page)
recolor('F39', GREEN)

# Memo items (from Model)
for col in ['G','H','I','J','K']:
    recolor(f'{col}56', GREEN)   # SBC
for col in ['F','G','H','I','J','K']:
    recolor(f'{col}57', GREEN)   # EBITDA

# P16: Implied TV/EBITDA (references Model tab)
recolor('P16', GREEN)

# ── BLACK: Local DCF formulas (already black, but be explicit) ──
local_formulas = ['P10','P11','P14']
for col in ['G','H','I','J','K']:
    local_formulas.extend([
        f'{col}10',  # Rev growth
        f'{col}13',  # EBIT margin
        f'{col}14',  # Taxes
        f'{col}15',  # NOPAT
        f'{col}18',  # NWC change ref
        f'{col}19',  # UFCF
        f'{col}21',  # Discount factor
        f'{col}22',  # Discounted CFs
    ])
local_formulas.extend(['F13',  # FY2025 EBIT margin
    'F24','F25','F26','F27','F28',  # Valuation
    'F31','F35','F37','F40',  # Equity bridge
])
# NWC totals and changes
for col in ['F','G','H','I','J','K']:
    local_formulas.append(f'{col}49')  # NWC total
for col in ['G','H','I','J','K']:
    local_formulas.append(f'{col}51')  # Change in NWC
for col in ['H','I','J','K']:
    local_formulas.append(f'{col}53')  # UFCF growth

for addr in local_formulas:
    c = ws[addr]
    if c.value is not None:
        recolor(addr, BLACK)

wb.save('NVIDIA NVDA US.xlsx')
print('Colors applied successfully!')
