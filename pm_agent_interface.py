"""
PM Agent Interface for NVIDIA DCF Model
========================================
Python interface for the Portfolio Manager agent to read/modify key drivers
in the NVIDIA analyst model and compute implied DCF valuations.

Key Drivers (8 levers):
  Revenue (5 segments): datacenter_growth, gaming_growth, automotive_growth,
                         proviz_growth, oem_growth
  Margins (3 inputs):   gm_improvement_bps, rd_improvement_bps, sga_improvement_bps

Usage:
    engine = NVDADCFEngine('NVIDIA NVDA US.xlsx')

    # View current drivers and valuation
    engine.print_summary()

    # Update drivers (partial dict — only specify what changes)
    engine.update_drivers({
        'datacenter_growth': {'FY2028': 0.30, 'FY2029': 0.25},
        'gm_improvement_bps': {'FY2028': 50, 'FY2029': 50},
    })

    # Recompute and display
    result = engine.compute_dcf()
    engine.print_summary()

    # Save updated drivers back to Excel
    engine.save()
"""

import openpyxl
from openpyxl.utils import column_index_from_string as _ci
from copy import deepcopy


# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════

# Model tab column letters for each period
_MODEL_QUARTERLY_COLS = {
    # FY2025 quarters (prior year base for Y/Y growth)
    'Q1-25': 'BO', 'Q2-25': 'BP', 'Q3-25': 'BQ', 'Q4-25': 'BR',
    # FY2026 quarters
    'Q1-26': 'BT', 'Q2-26': 'BU', 'Q3-26': 'BV', 'Q4-26': 'BW',
    # FY2027 quarters
    'Q1-27': 'BY', 'Q2-27': 'BZ', 'Q3-27': 'CA', 'Q4-27': 'CB',
}

_MODEL_ANNUAL_COLS = {
    'FY2025': 'BS', 'FY2026': 'BX', 'FY2027': 'CC',
    'FY2028': 'CD', 'FY2029': 'CE', 'FY2030': 'CF',
}

# Periods the PM agent can modify (Q4-26 onward are estimates)
_DRIVER_PERIODS_QTR = ['Q4-26', 'Q1-27', 'Q2-27', 'Q3-27', 'Q4-27']
_DRIVER_PERIODS_ANN = ['FY2028', 'FY2029', 'FY2030']
_DRIVER_PERIODS_ALL = _DRIVER_PERIODS_QTR + _DRIVER_PERIODS_ANN

# Segment definitions: name -> (revenue_row, growth_row)
_SEGMENTS = {
    'datacenter':  (22, 24),
    'gaming':      (7,  9),
    'automotive':  (26, 28),
    'proviz':      (11, 13),
    'oem':         (30, 32),
}

# Margin driver definitions: name -> (margin_%_row, bps_row, sign)
# sign: +1 means higher bps = higher margin (gross margin)
#       -1 means higher bps = lower margin % (R&D, SG&A — improvement = less spend)
_MARGIN_DRIVERS = {
    'gm_improvement_bps':  (59, 62, +1),
    'rd_improvement_bps':  (76, 77, -1),
    'sga_improvement_bps': (91, 92, -1),
}

# Other Model tab rows
_ROW_TOTAL_REV = 34
_ROW_EBIT = 386
_ROW_EBIT_MARGIN = 109
_ROW_DA = 933
_ROW_CAPEX = 934
_ROW_AR = 954
_ROW_INV = 958
_ROW_AP = 985
_ROW_ACCRUED = 987
_ROW_CASH = 952
_ROW_MKT_SEC = 953
_ROW_LT_DEBT = 992
_ROW_ST_DEBT = 988
_ROW_SHARES = 429

_PROJ_FY = ['FY2026', 'FY2027', 'FY2028', 'FY2029', 'FY2030']


# ═══════════════════════════════════════════════════════════════
# ENGINE
# ═══════════════════════════════════════════════════════════════

class NVDADCFEngine:
    """Read/modify NVIDIA DCF drivers and compute implied valuations."""

    def __init__(self, filepath):
        self.filepath = filepath
        print(f"    [DCF] Loading Excel model from {filepath}...")
        self._load_baseline()
        self.drivers = self.get_drivers()
        self._last_result = None
        print(f"    [DCF] Model loaded — {len(self.drivers)} drivers, "
              f"stock price: ${self._stock_price:,.2f}, "
              f"WACC assumptions: rf={self._dcf_assumptions['rf']:.2%}, "
              f"beta={self._dcf_assumptions['beta']:.2f}")

    # ───────────────────────────────────────────────────────────
    # LOADING
    # ───────────────────────────────────────────────────────────

    def _read_cell(self, ws, row, col_letter):
        """Read a cell value, returning 0.0 for None."""
        v = ws.cell(row=row, column=_ci(col_letter)).value
        return float(v) if v is not None else 0.0

    def _load_baseline(self):
        """Load all baseline data from the Excel file."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb['Model']

        # -- Quarterly segment revenues (FY2025 Q1-Q4 and FY2026 Q1-Q3) --
        self._q_rev = {}  # {segment: {period: value}}
        q_periods_to_read = {
            **{k: v for k, v in _MODEL_QUARTERLY_COLS.items()
               if k.endswith('-25') or (k.endswith('-26') and not k.startswith('Q4'))},
        }
        for seg, (rev_row, _) in _SEGMENTS.items():
            self._q_rev[seg] = {}
            for period, col in q_periods_to_read.items():
                self._q_rev[seg][period] = self._read_cell(ws, rev_row, col)

        # -- Annual data --
        self._annual = {}
        for fy, col in _MODEL_ANNUAL_COLS.items():
            self._annual[fy] = {
                'total_rev': self._read_cell(ws, _ROW_TOTAL_REV, col),
                'ebit':      self._read_cell(ws, _ROW_EBIT, col),
                'ebit_margin': self._read_cell(ws, _ROW_EBIT_MARGIN, col),
                'da':        self._read_cell(ws, _ROW_DA, col),
                'capex':     self._read_cell(ws, _ROW_CAPEX, col),
                'ar':        self._read_cell(ws, _ROW_AR, col),
                'inv':       self._read_cell(ws, _ROW_INV, col),
                'ap':        self._read_cell(ws, _ROW_AP, col),
                'accrued':   self._read_cell(ws, _ROW_ACCRUED, col),
            }
            # Margins
            for drv, (margin_row, _, _) in _MARGIN_DRIVERS.items():
                key = drv.replace('_improvement_bps', '_pct')
                self._annual[fy][key] = self._read_cell(ws, margin_row, col)

        # -- Q3-26 margins (baseline for quarterly cascade) --
        bv = 'BV'
        self._q3_26_margins = {}
        for drv, (margin_row, _, _) in _MARGIN_DRIVERS.items():
            key = drv.replace('_improvement_bps', '_pct')
            self._q3_26_margins[key] = self._read_cell(ws, margin_row, bv)

        # -- Balance sheet (FY2025) --
        bs = _MODEL_ANNUAL_COLS['FY2025']
        self._bs = {
            'cash':    self._read_cell(ws, _ROW_CASH, bs),
            'mkt_sec': self._read_cell(ws, _ROW_MKT_SEC, bs),
            'lt_debt': self._read_cell(ws, _ROW_LT_DEBT, bs),
            'st_debt': self._read_cell(ws, _ROW_ST_DEBT, bs),
            'shares':  self._read_cell(ws, _ROW_SHARES, bs),
        }

        # -- Stock price (DO NOT use Excel - it's stale) --
        # Will fetch live price from stock market agent when compute_dcf() is called
        self._stock_price = None

        # -- DCF assumptions (from DCF tab) --
        ws_dcf = wb['DCF']
        self._dcf_assumptions = {
            'rf':   float(ws_dcf['P6'].value or 0.0425),
            'erp':  float(ws_dcf['P7'].value or 0.055),
            'beta': float(ws_dcf['P8'].value or 1.65),
            'ev':   float(ws_dcf['P9'].value or 0.998),
            'tax':  float(ws_dcf['P12'].value or 0.15),
            'cod':  float(ws_dcf['P13'].value or 0.035),
            'tgr':  float(ws_dcf['P15'].value or 0.03),
        }

        wb.close()

    # ───────────────────────────────────────────────────────────
    # DRIVERS: GET / UPDATE
    # ───────────────────────────────────────────────────────────

    def get_drivers(self):
        """
        Read current driver values from the Excel file.

        Returns dict with keys:
          datacenter_growth, gaming_growth, automotive_growth,
          proviz_growth, oem_growth,
          gm_improvement_bps, rd_improvement_bps, sga_improvement_bps

        Each value is a dict mapping period -> value, e.g.:
          {'Q4-26': 0.67, 'Q1-27': 0.50, ..., 'FY2028': 0.25, ...}
        """
        wb = openpyxl.load_workbook(self.filepath, data_only=False)
        ws = wb['Model']
        drivers = {}

        # Revenue growth by segment
        for seg, (_, growth_row) in _SEGMENTS.items():
            key = f'{seg}_growth'
            drivers[key] = {}
            for period in _DRIVER_PERIODS_ALL:
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    v = ws.cell(row=growth_row, column=_ci(col)).value
                    # Only include if it's a hardcoded number (not a formula)
                    if v is not None and not (isinstance(v, str) and v.startswith('=')):
                        drivers[key][period] = float(v)

        # Margin bps drivers
        for drv, (_, bps_row, _) in _MARGIN_DRIVERS.items():
            drivers[drv] = {}
            for period in _DRIVER_PERIODS_ALL:
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    v = ws.cell(row=bps_row, column=_ci(col)).value
                    if v is not None and not (isinstance(v, str) and v.startswith('=')):
                        drivers[drv][period] = float(v)

        wb.close()
        return drivers

    def update_drivers(self, changes):
        """
        Update driver values. Accepts a partial dict — only specify what changes.

        Args:
            changes: dict matching the structure of get_drivers(), e.g.:
                {
                    'datacenter_growth': {'FY2028': 0.30, 'FY2029': 0.25},
                    'gm_improvement_bps': {'FY2028': 50},
                }
        """
        for key, period_vals in changes.items():
            if key not in self.drivers:
                raise KeyError(f"Unknown driver: {key}. Valid: {list(self.drivers.keys())}")
            for period, val in period_vals.items():
                if period not in _DRIVER_PERIODS_ALL:
                    raise KeyError(f"Invalid period: {period}. Valid: {_DRIVER_PERIODS_ALL}")
                self.drivers[key][period] = float(val)

    # ───────────────────────────────────────────────────────────
    # COMPUTE
    # ───────────────────────────────────────────────────────────

    def compute_dcf(self, current_price: float = None):
        """
        Run the full DCF calculation with current drivers.

        Args:
            current_price: Optional current stock price. If not provided, will fetch live from market.

        Returns a dict with:
          revenue, ebit, ebit_margin, ufcf (per FY),
          sum_pv_ufcf, terminal_value, pv_terminal_value,
          enterprise_value, equity_value, implied_price,
          current_price, upside, wacc
        """
        # Fetch live stock price if not provided
        if current_price is None:
            try:
                from agents.stock_market_agent import StockMarketAgent
                agent = StockMarketAgent()
                current_price = agent.get_price('NVDA', use_cache=True)
                if current_price <= 0:
                    current_price = self._stock_price or 100.0  # Fallback
            except Exception:
                current_price = self._stock_price or 100.0  # Fallback if fetch fails

        self._stock_price = current_price
        drivers = self.drivers

        print(f"    [DCF] Step 1/7: Building segment revenues...")
        # ── Step 1: Compute revenues by segment ──────────────
        seg_rev = {}   # {segment: {FY: value}}
        total_rev = {} # {FY: value}

        for seg, (_, _) in _SEGMENTS.items():
            seg_rev[seg] = {}
            growth_key = f'{seg}_growth'
            growth = drivers.get(growth_key, {})

            # FY2025: sum of known quarterly actuals
            seg_rev[seg]['FY2025'] = sum(
                self._q_rev[seg].get(f'Q{q}-25', 0) for q in [1,2,3,4]
            )

            # FY2026: Q1-Q3 actuals + Q4 estimated
            q4_25 = self._q_rev[seg].get('Q4-25', 0)
            q4_26_growth = growth.get('Q4-26', 0)
            q4_26 = q4_25 * (1 + q4_26_growth)
            seg_rev[seg]['FY2026'] = (
                self._q_rev[seg].get('Q1-26', 0) +
                self._q_rev[seg].get('Q2-26', 0) +
                self._q_rev[seg].get('Q3-26', 0) +
                q4_26
            )

            # FY2027: each quarter = same quarter FY2026 × (1 + Y/Y growth)
            q26_vals = {
                'Q1': self._q_rev[seg].get('Q1-26', 0),
                'Q2': self._q_rev[seg].get('Q2-26', 0),
                'Q3': self._q_rev[seg].get('Q3-26', 0),
                'Q4': q4_26,
            }
            fy27_total = 0
            for q in ['Q1', 'Q2', 'Q3', 'Q4']:
                qg = growth.get(f'{q}-27', 0)
                fy27_total += q26_vals[q] * (1 + qg)
            seg_rev[seg]['FY2027'] = fy27_total

            # FY2028-FY2030: annual growth
            for fy in ['FY2028', 'FY2029', 'FY2030']:
                prev_fy = f'FY{int(fy[2:]) - 1}'
                g = growth.get(fy, 0)
                seg_rev[seg][fy] = seg_rev[seg][prev_fy] * (1 + g)

        # Total revenue = sum of segments
        for fy in ['FY2025'] + _PROJ_FY:
            total_rev[fy] = sum(seg_rev[seg][fy] for seg in _SEGMENTS)

        print(f"    [DCF] Step 2/7: Computing margin cascade...")
        # ── Step 2: Compute margins ──────────────────────────
        # Quarterly margin cascade from Q3-26 baseline
        gm_q = {'Q3-26': self._q3_26_margins.get('gm_pct', 0.736)}
        rd_q = {'Q3-26': self._q3_26_margins.get('rd_pct', 0.061)}
        sga_q = {'Q3-26': self._q3_26_margins.get('sga_pct', 0.013)}

        prev_q = 'Q3-26'
        q_periods_ordered = ['Q4-26', 'Q1-27', 'Q2-27', 'Q3-27', 'Q4-27']
        for qp in q_periods_ordered:
            gm_bps = drivers.get('gm_improvement_bps', {}).get(qp, 0)
            rd_bps = drivers.get('rd_improvement_bps', {}).get(qp, 0)
            sga_bps = drivers.get('sga_improvement_bps', {}).get(qp, 0)
            gm_q[qp] = gm_q[prev_q] + gm_bps / 10000
            rd_q[qp] = rd_q[prev_q] - rd_bps / 10000
            sga_q[qp] = sga_q[prev_q] - sga_bps / 10000
            prev_q = qp

        # Compute FY2026 and FY2027 annual margins (revenue-weighted)
        margins = {}

        # FY2025: use baseline
        margins['FY2025'] = {
            'gm': self._annual['FY2025'].get('gm_pct', 0.755),
            'rd': self._annual['FY2025'].get('rd_pct', 0.072),
            'sga': self._annual['FY2025'].get('sga_pct', 0.018),
        }

        # FY2026: Q1-Q3 use baseline annual, Q4 uses cascade
        # Simplification: weight Q1-Q3 at baseline annual margin, Q4 at cascaded margin
        fy26_rev = total_rev['FY2026']
        q123_rev_26 = sum(
            sum(self._q_rev[seg].get(f'Q{q}-26', 0) for seg in _SEGMENTS)
            for q in [1, 2, 3]
        )
        q4_rev_26 = fy26_rev - q123_rev_26
        baseline_gm_26 = self._annual['FY2026'].get('gm_pct', 0.712)
        baseline_rd_26 = self._annual['FY2026'].get('rd_pct', 0.064)
        baseline_sga_26 = self._annual['FY2026'].get('sga_pct', 0.014)

        if fy26_rev > 0:
            # Revenue-weighted: Q1-Q3 at baseline, Q4 at cascade
            margins['FY2026'] = {
                'gm':  (q123_rev_26 * baseline_gm_26 + q4_rev_26 * gm_q['Q4-26']) / fy26_rev,
                'rd':  (q123_rev_26 * baseline_rd_26 + q4_rev_26 * rd_q['Q4-26']) / fy26_rev,
                'sga': (q123_rev_26 * baseline_sga_26 + q4_rev_26 * sga_q['Q4-26']) / fy26_rev,
            }
        else:
            margins['FY2026'] = {'gm': baseline_gm_26, 'rd': baseline_rd_26, 'sga': baseline_sga_26}

        # FY2027: revenue-weighted quarterly margins
        fy27_q_revs = {}
        for seg in _SEGMENTS:
            q26_vals = {
                'Q1': self._q_rev[seg].get('Q1-26', 0),
                'Q2': self._q_rev[seg].get('Q2-26', 0),
                'Q3': self._q_rev[seg].get('Q3-26', 0),
                'Q4': self._q_rev[seg].get('Q4-25', 0) * (1 + drivers.get(f'{seg}_growth', {}).get('Q4-26', 0)),
            }
            for q_num, q_name in enumerate(['Q1', 'Q2', 'Q3', 'Q4'], 1):
                period = f'{q_name}-27'
                qg = drivers.get(f'{seg}_growth', {}).get(period, 0)
                rev = q26_vals[q_name] * (1 + qg)
                fy27_q_revs[period] = fy27_q_revs.get(period, 0) + rev

        fy27_rev = total_rev['FY2027']
        if fy27_rev > 0:
            margins['FY2027'] = {
                'gm':  sum(fy27_q_revs.get(f'Q{q}-27', 0) * gm_q[f'Q{q}-27'] for q in [1,2,3,4]) / fy27_rev,
                'rd':  sum(fy27_q_revs.get(f'Q{q}-27', 0) * rd_q[f'Q{q}-27'] for q in [1,2,3,4]) / fy27_rev,
                'sga': sum(fy27_q_revs.get(f'Q{q}-27', 0) * sga_q[f'Q{q}-27'] for q in [1,2,3,4]) / fy27_rev,
            }
        else:
            margins['FY2027'] = {'gm': gm_q['Q4-27'], 'rd': rd_q['Q4-27'], 'sga': sga_q['Q4-27']}

        # FY2028-FY2030: cascade from prior year annual margin
        for fy in ['FY2028', 'FY2029', 'FY2030']:
            prev_fy = f'FY{int(fy[2:]) - 1}'
            gm_bps = drivers.get('gm_improvement_bps', {}).get(fy, 0)
            rd_bps = drivers.get('rd_improvement_bps', {}).get(fy, 0)
            sga_bps = drivers.get('sga_improvement_bps', {}).get(fy, 0)
            margins[fy] = {
                'gm':  margins[prev_fy]['gm']  + gm_bps / 10000,
                'rd':  margins[prev_fy]['rd']  - rd_bps / 10000,
                'sga': margins[prev_fy]['sga'] - sga_bps / 10000,
            }

        print(f"    [DCF] Step 3/7: Computing EBIT...")
        # ── Step 3: EBIT ─────────────────────────────────────
        ebit = {}
        ebit_margin = {}
        for fy in _PROJ_FY:
            m = margins[fy]
            ebit_margin[fy] = m['gm'] - m['rd'] - m['sga']
            ebit[fy] = total_rev[fy] * ebit_margin[fy]

        print(f"    [DCF] Step 4/7: Computing UFCF (unlevered free cash flow)...")
        # ── Step 4: UFCF ─────────────────────────────────────
        a = self._dcf_assumptions
        tax_rate = a['tax']

        # NWC from baseline (scale with revenue)
        nwc = {}
        for fy in ['FY2025'] + _PROJ_FY:
            bl = self._annual[fy]
            bl_rev = bl['total_rev']
            new_rev = total_rev[fy]
            scale = new_rev / bl_rev if bl_rev > 0 else 1.0
            nwc[fy] = (bl['ar'] + bl['inv'] - bl['ap'] - bl['accrued']) * scale

        ufcf = {}
        for fy in _PROJ_FY:
            nopat = ebit[fy] * (1 - tax_rate)
            da = self._annual[fy]['da']
            capex = self._annual[fy]['capex']
            prev_fy = f'FY{int(fy[2:]) - 1}'
            delta_nwc = nwc[fy] - nwc[prev_fy]
            ufcf[fy] = nopat + da - capex - delta_nwc

        print(f"    [DCF] Step 5/7: Computing WACC...")
        # ── Step 5: WACC ─────────────────────────────────────
        ev = a['ev']
        dv = 1 - ev
        lev_beta = a['beta'] * (1 + (1 - a['tax']) * dv / ev) if ev > 0 else a['beta']
        coe = a['rf'] + a['erp'] * lev_beta
        wacc = coe * ev + a['cod'] * (1 - a['tax']) * dv

        print(f"    [DCF] Step 6/7: Discounting cash flows & terminal value...")
        # ── Step 6: Discount & Terminal Value ─────────────────
        pv_ufcf = {}
        for i, fy in enumerate(_PROJ_FY, 1):
            df = 1 / (1 + wacc) ** i
            pv_ufcf[fy] = ufcf[fy] * df

        sum_pv = sum(pv_ufcf.values())

        last_ufcf = ufcf['FY2030']
        tv = last_ufcf * (1 + a['tgr']) / (wacc - a['tgr'])
        df_tv = 1 / (1 + wacc) ** len(_PROJ_FY)
        pv_tv = tv * df_tv

        tev = sum_pv + pv_tv

        print(f"    [DCF] Step 7/7: Equity bridge (TEV -> implied price)...")
        # ── Step 7: Equity Bridge ─────────────────────────────
        net_debt = self._bs['lt_debt'] + self._bs['st_debt'] - self._bs['cash'] - self._bs['mkt_sec']
        equity_value = tev - net_debt
        implied_price = equity_value / self._bs['shares'] if self._bs['shares'] > 0 else 0

        print(f"    [DCF] Done — implied price: ${implied_price:,.2f} "
              f"(WACC: {wacc:.1%}, TEV: ${tev:,.0f})")
        # ── Build result ──────────────────────────────────────
        self._last_result = {
            'seg_rev': seg_rev,
            'total_rev': total_rev,
            'margins': margins,
            'ebit': ebit,
            'ebit_margin': ebit_margin,
            'ufcf': ufcf,
            'pv_ufcf': pv_ufcf,
            'nwc': nwc,
            'sum_pv_ufcf': sum_pv,
            'terminal_value': tv,
            'pv_terminal_value': pv_tv,
            'enterprise_value': tev,
            'net_debt': net_debt,
            'equity_value': equity_value,
            'implied_price': implied_price,
            'current_price': self._stock_price,
            'upside': (implied_price / self._stock_price - 1) if self._stock_price > 0 else 0,
            'wacc': wacc,
            'shares': self._bs['shares'],
        }
        return self._last_result

    # ───────────────────────────────────────────────────────────
    # SAVE TO EXCEL
    # ───────────────────────────────────────────────────────────

    def save(self):
        """Write current driver values back to the Model tab in Excel."""
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb['Model']

        # Write segment growth rates
        for seg, (_, growth_row) in _SEGMENTS.items():
            growth_key = f'{seg}_growth'
            for period, val in self.drivers.get(growth_key, {}).items():
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    ws.cell(row=growth_row, column=_ci(col), value=val)

        # Write margin bps drivers
        for drv, (_, bps_row, _) in _MARGIN_DRIVERS.items():
            for period, val in self.drivers.get(drv, {}).items():
                col = _MODEL_QUARTERLY_COLS.get(period) or _MODEL_ANNUAL_COLS.get(period)
                if col:
                    ws.cell(row=bps_row, column=_ci(col), value=val)

        wb.save(self.filepath)
        print(f'Drivers saved to {self.filepath}')

    # ───────────────────────────────────────────────────────────
    # DISPLAY
    # ───────────────────────────────────────────────────────────

    def print_drivers(self):
        """Print all current driver values in a readable table."""
        print('\n' + '=' * 88)
        print('  CURRENT KEY DRIVERS')
        print('=' * 88)

        # Revenue growth
        print('\n  Revenue Growth Rates:')
        header = f"  {'Segment':<18}"
        for p in _DRIVER_PERIODS_ALL:
            header += f'{p:>8}'
        print(header)
        print('  ' + '-' * 84)

        for seg in _SEGMENTS:
            key = f'{seg}_growth'
            row = f'  {seg.upper():<18}'
            for p in _DRIVER_PERIODS_ALL:
                v = self.drivers.get(key, {}).get(p)
                row += f'{v:>7.0%} ' if v is not None else f'{"N/A":>8}'
            print(row)

        # Margin improvements
        print('\n  Margin Improvements (bps):')
        header = f"  {'Driver':<18}"
        for p in _DRIVER_PERIODS_ALL:
            header += f'{p:>8}'
        print(header)
        print('  ' + '-' * 84)

        for drv in _MARGIN_DRIVERS:
            label = drv.replace('_improvement_bps', '').upper()
            row = f'  {label:<18}'
            for p in _DRIVER_PERIODS_ALL:
                v = self.drivers.get(drv, {}).get(p)
                row += f'{v:>7.0f} ' if v is not None else f'{"N/A":>8}'
            print(row)
        print()

    def print_summary(self):
        """Print drivers and valuation summary."""
        if self._last_result is None:
            self.compute_dcf()
        r = self._last_result

        self.print_drivers()

        print('=' * 88)
        print('  VALUATION SUMMARY')
        print('=' * 88)

        # Revenue and EBIT table
        header = f"\n  {'($ in millions)':<34}"
        for fy in _PROJ_FY:
            header += f'{fy:>12}'
        print(header)
        print('  ' + '-' * 84)

        # Segment revenues
        for seg in _SEGMENTS:
            label = seg.capitalize()
            row = f'  {label + " Revenue":<34}'
            for fy in _PROJ_FY:
                row += f'{r["seg_rev"][seg][fy]:>12,.0f}'
            print(row)

        row = f'  {"TOTAL REVENUE":<34}'
        for fy in _PROJ_FY:
            row += f'{r["total_rev"][fy]:>12,.0f}'
        print(row)

        row = f'  {"  Y/Y Growth":<34}'
        for fy in _PROJ_FY:
            prev = f'FY{int(fy[2:])-1}'
            g = r['total_rev'][fy] / r['total_rev'][prev] - 1 if r['total_rev'].get(prev) else 0
            row += f'{g:>11.1%} '
        print(row)

        print()
        for label, mkey in [('Gross Margin %', 'gm'), ('R&D Margin %', 'rd'),
                             ('SG&A Margin %', 'sga'), ('EBIT Margin %', None)]:
            row = f'  {label:<34}'
            for fy in _PROJ_FY:
                if mkey:
                    row += f'{r["margins"][fy][mkey]:>11.1%} '
                else:
                    row += f'{r["ebit_margin"][fy]:>11.1%} '
            print(row)

        row = f'  {"NON-GAAP EBIT":<34}'
        for fy in _PROJ_FY:
            row += f'{r["ebit"][fy]:>12,.0f}'
        print(row)

        print()
        row = f'  {"UFCF":<34}'
        for fy in _PROJ_FY:
            row += f'{r["ufcf"][fy]:>12,.0f}'
        print(row)

        row = f'  {"PV of UFCF":<34}'
        for fy in _PROJ_FY:
            row += f'{r["pv_ufcf"][fy]:>12,.0f}'
        print(row)

        # Valuation
        print(f'\n  {"─" * 50}')
        print(f'  WACC:                         {r["wacc"]:.1%}')
        print(f'  Sum of PV(UFCF):              ${r["sum_pv_ufcf"]:>14,.0f}')
        print(f'  Terminal Value:                ${r["terminal_value"]:>14,.0f}')
        print(f'  PV of Terminal Value:          ${r["pv_terminal_value"]:>14,.0f}')
        print(f'  Enterprise Value:              ${r["enterprise_value"]:>14,.0f}')
        print(f'  Net Debt (Cash):               ${r["net_debt"]:>14,.0f}')
        print(f'  Equity Value:                  ${r["equity_value"]:>14,.0f}')
        print(f'  Shares Outstanding (mm):       {r["shares"]:>14,.0f}')
        print(f'  {"─" * 50}')
        print(f'  IMPLIED SHARE PRICE:           ${r["implied_price"]:>13,.2f}')
        print(f'  Current Price:                 ${r["current_price"]:>13,.2f}')

        upside = r['upside']
        direction = 'UPSIDE' if upside >= 0 else 'DOWNSIDE'
        print(f'  {direction}:                     {upside:>12.1%}')
        print(f'  {"─" * 50}\n')


# ═══════════════════════════════════════════════════════════════
# CONVENIENCE FUNCTION
# ═══════════════════════════════════════════════════════════════

def quick_valuation(filepath, changes=None):
    """
    One-liner: load model, optionally apply changes, print summary.

    Args:
        filepath: path to NVIDIA NVDA US.xlsx
        changes: optional driver changes dict (same format as update_drivers)

    Returns:
        NVDADCFEngine instance (for further interaction)
    """
    engine = NVDADCFEngine(filepath)
    if changes:
        engine.update_drivers(changes)
    engine.compute_dcf()
    engine.print_summary()
    return engine


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'NVIDIA NVDA US.xlsx'

    print('Loading baseline model...')
    engine = NVDADCFEngine(path)
    engine.compute_dcf()
    engine.print_summary()

    # Example: PM agent scenario — bullish datacenter view
    print('\n' + '=' * 88)
    print('  SCENARIO: Bullish Datacenter (PM agent applies council view)')
    print('=' * 88)

    engine.update_drivers({
        'datacenter_growth': {'FY2028': 0.30, 'FY2029': 0.25, 'FY2030': 0.22},
        'gm_improvement_bps': {'FY2028': 50, 'FY2029': 50, 'FY2030': 50},
    })
    engine.compute_dcf()
    engine.print_summary()
