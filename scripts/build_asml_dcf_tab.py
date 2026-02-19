"""
build_asml_dcf_tab.py
=====================
Builds the full DCF tab in the ASML analyst model Excel file.
ASML is a semiconductor equipment manufacturer with product-based revenue.

Key characteristics:
- Revenue from equipment sales: EUV, ArFi, ArF, KrF, i-line, Metrology
- Pricing power from EUV dominance (~40% of revenue)
- Gross margin >50% (equipment mix)
- Capex modest relative to foundries/designers

Usage:
    cd agent-alpha
    python scripts/build_asml_dcf_tab.py
"""

import openpyxl
from openpyxl.styles import Font

filepath = 'financial_models/ASML Holding ASML NA.xlsx'
wb = openpyxl.load_workbook(filepath)

if 'DCF' in wb.sheetnames:
    idx = wb.sheetnames.index('DCF')
    del wb['DCF']
else:
    idx = 1
ws = wb.create_sheet('DCF', idx)

# ── Column mappings (ASML uses different column structure) ─────
MC = {'FY2025': 'L', 'FY2026': 'Q', 'FY2027': 'V', 'FY2028': 'AA', 'FY2029': 'AF'}
DC = {'FY2025': 'F', 'FY2026': 'G', 'FY2027': 'H', 'FY2028': 'I', 'FY2029': 'J'}
ALL = ['FY2025', 'FY2026', 'FY2027', 'FY2028', 'FY2029']
PROJ = ['FY2026', 'FY2027', 'FY2028', 'FY2029']

# ── Styles ───────────────────────────────────────────────────
f_title = Font(color='000000', bold=True, size=14)
f_section = Font(color='000000', bold=True, size=12)
f_green = Font(color='007F00', size=10)

# ── Setup ────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 28
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[col].width = 12

ws['A1'] = "ASML DCF Valuation Analysis"
ws['A1'].font = f_title

# ── Headers (Row 4) ──────────────────────────────────────────
row = 4
ws[f'A{row}'] = "Period"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = period
    ws[f'{col}{row}'].font = Font(bold=True)

# ── Revenue Build (Rows 6-15) ────────────────────────────────
# ASML has: EUV + ArFi + ArF + KrF + i-line + Metrology = Total
row = 6
ws[f'A{row}'] = "REVENUE BUILD"
ws[f'A{row}'].font = f_section

equipment_types = [
    ('EUV', 45),       # Row 41 = Total EUV (from units × price)
    ('ArFi', 48),      # Row 48 = Total ArFi
    ('ArF', 51),       # Row 51 = Total ArF
    ('KrF', 54),       # Row 54 = Total KrF
    ('i-line', 57),    # Row 57 = Total i-line
    ('Metrology', 60), # Row 60 = Total Metrology
]

row = 7
for eq_type, model_row in equipment_types:
    ws[f'A{row}'] = f"{eq_type} Revenue"
    for period in ALL:
        col = DC[period]
        model_col = MC[period]
        ws[f'{col}{row}'] = f"=Model!{model_col}{model_row}"
        ws[f'{col}{row}'].font = f_green
    row += 1

# Total Revenue (Row 13)
ws['A13'] = "Total Revenue"
ws['A13'].font = Font(bold=True)
for period in ALL:
    col = DC[period]
    ws[f'{col}13'] = f"=SUM({col}7:{col}12)"

# Y/Y Growth (Row 14)
row = 14
ws[f'A{row}'] = "Y/Y Growth %"
for period in PROJ:
    col = DC[period]
    prev_col = DC[ALL[ALL.index(period)-1]]
    ws[f'{col}{row}'] = f"={col}13/{prev_col}13 - 1"

# ── Operating Margins (Rows 16-19) ────────────────────────────
row = 16
ws[f'A{row}'] = "OPERATING MARGINS"
ws[f'A{row}'].font = f_section

row = 17
ws[f'A{row}'] = "Gross Margin %"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}74"  # Row 74 = GM%
    ws[f'{col}{row}'].font = f_green

row = 18
ws[f'A{row}'] = "Operating Margin %"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}89/Model!{model_col}67"  # EBIT / Revenue
    ws[f'{col}{row}'].font = f_green

# ── UFCF Build (Rows 21-31) ──────────────────────────────────
row = 21
ws[f'A{row}'] = "UFCF BUILD"
ws[f'A{row}'].font = f_section

row = 22
ws[f'A{row}'] = "Revenue"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}13"

row = 23
ws[f'A{row}'] = "EBIT"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}89"
    ws[f'{col}{row}'].font = f_green

row = 24
ws[f'A{row}'] = "Tax Rate %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = 0.14  # ASML effective tax ~14%
    ws[f'{col}{row}'].number_format = '0.0%'

row = 25
ws[f'A{row}'] = "NOPAT"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}23 * (1 - {col}24)"

row = 26
ws[f'A{row}'] = "D&A"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}113"  # Row 113 = D&A
    ws[f'{col}{row}'].font = f_green

row = 27
ws[f'A{row}'] = "CapEx"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}22 * 0.03"  # 3% of revenue (asset-light business)

row = 28
ws[f'A{row}'] = "Change in Working Capital"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}22 * 0.05"  # 5% of revenue

row = 29
ws[f'A{row}'] = "Unlevered FCF"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}25 + {col}26 - {col}27 - {col}28"

# ── Discounting (Rows 31-35) ──────────────────────────────────
row = 31
ws[f'A{row}'] = "DISCOUNTING"
ws[f'A{row}'].font = f_section

row = 32
ws[f'A{row}'] = "WACC %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = 0.085  # ASML WACC: ~8.5% (defensive, high margin)
    ws[f'{col}{row}'].number_format = '0.0%'

row = 33
ws[f'A{row}'] = "Discount Factor"
for i, period in enumerate(ALL, 1):
    col = DC[period]
    ws[f'{col}{row}'] = f"=1 / (1 + {col}32)^{i}"

row = 34
ws[f'A{row}'] = "PV of UFCF"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}29 * {col}33"

# ── Terminal Value (Rows 36-38) ──────────────────────────────
row = 36
ws[f'A{row}'] = "TERMINAL VALUE"
ws[f'A{row}'].font = f_section

row = 37
ws[f'A{row}'] = "Terminal Growth %"
ws['F37'] = 0.02  # 2% perpetual growth
ws['F37'].number_format = '0.0%'

row = 38
ws[f'A{row}'] = "Terminal Value"
ws['F38'] = f"=F29 * (1 + F37) / (F32 - F37)"

row = 39
ws[f'A{row}'] = "PV of TV"
ws['F39'] = f"=F38 * F33"

# ── Valuation (Rows 41-48) ───────────────────────────────────
row = 41
ws[f'A{row}'] = "VALUATION"
ws[f'A{row}'].font = f_section

row = 42
ws[f'A{row}'] = "Sum PV UFCF"
ws['F42'] = f"=SUM(F34:J34)"

row = 43
ws[f'A{row}'] = "PV Terminal Value"
ws['F43'] = f"=F39"

row = 44
ws[f'A{row}'] = "Enterprise Value"
ws['F44'] = f"=F42 + F43"

row = 45
ws[f'A{row}'] = "Less: Net Debt"
ws['F45'] = f"=Model!L700 - Model!L650"  # Adjust rows as needed

row = 46
ws[f'A{row}'] = "Equity Value"
ws['F46'] = f"=F44 - F45"

row = 47
ws[f'A{row}'] = "Shares Outstanding (mm)"
ws['F47'] = f"=Model!L750"  # Adjust row

row = 48
ws[f'A{row}'] = "Implied Price"
ws['F48'] = f"=F46 / F47"

row = 49
ws[f'A{row}'] = "Current Price"
ws['F49'] = f"='Front Page'!H20"

row = 50
ws[f'A{row}'] = "Upside %"
ws['F50'] = f"=F48 / F49 - 1"
ws['F50'].number_format = '0.0%'

wb.save(filepath)
print(f"✅ DCF tab built for ASML!")
print(f"   Rows: 50, Cols: 10")
print(f"   Revenue: 6 equipment types (EUV, ArFi, ArF, KrF, i-line, Metrology)")
print(f"   Projection periods: FY2025-FY2029")
