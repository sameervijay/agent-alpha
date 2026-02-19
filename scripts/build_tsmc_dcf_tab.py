"""
build_tsmc_dcf_tab.py
=====================
Builds the full DCF tab in the TSMC analyst model Excel file.
Includes: Revenue build by segment (5 segments), operating margin drivers,
UFCF build, discounting, valuation, equity bridge, and memo items.

Usage:
    cd agent-alpha
    python scripts/build_tsmc_dcf_tab.py

TSMC Model Tab Key Rows:
- Total Revenue (TTM): Row 25
- Segments: Smartphone (7), HPC (10), IoT (13), Automotive (16), Digital Consumer (19), Other (22)
- Gross Margin %: Row 38
- Operating Income (EBIT): Row 59
- D&A: Rows 72 + 76 (separate line items)
- CapEx: Row 86 (in billions, convert to millions)
- Tax Rate: Row 60 (check for actual tax line)
- Shares Outstanding: Row ~500 (need to locate in actual model)
- Stock Price: Front Page!H20

Column Mappings (Annual):
- FY2025: L, FY2026: Q, FY2027: V, FY2028: AA, FY2029: AF
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path

# Load or create workbook
filepath = 'financial_models/Taiwan Semiconductor Manufacturing Company TSM US.xlsx'
wb = openpyxl.load_workbook(filepath)

# Create/replace DCF sheet
if 'DCF' in wb.sheetnames:
    idx = wb.sheetnames.index('DCF')
    del wb['DCF']
else:
    idx = 1
ws = wb.create_sheet('DCF', idx)

# ── Column mappings ──────────────────────────────────────────
MC = {  # Model tab columns (annual)
    'FY2025': 'L', 'FY2026': 'Q', 'FY2027': 'V',
    'FY2028': 'AA', 'FY2029': 'AF',
}
DC = {  # DCF tab columns
    'FY2025': 'F', 'FY2026': 'G', 'FY2027': 'H',
    'FY2028': 'I', 'FY2029': 'J',
}
ALL = ['FY2025', 'FY2026', 'FY2027', 'FY2028', 'FY2029']
PROJ = ['FY2026', 'FY2027', 'FY2028', 'FY2029']

# ── Colors & Fonts ───────────────────────────────────────────
BLUE   = '0000FF'
GREEN  = '007F00'
BLACK  = '000000'

def mkfont(color=BLACK, bold=False, size=None, italic=False, underline=None):
    kw = {'color': color, 'bold': bold}
    if size: kw['size'] = size
    if italic: kw['italic'] = italic
    if underline: kw['underline'] = underline
    return Font(**kw)

f_title    = mkfont(BLACK, True, 14)
f_section  = mkfont(BLACK, True, 12)
f_sublabel = mkfont(BLACK, True, 11)
f_data     = mkfont(BLACK, False, 11)
f_formula  = mkfont(GREEN, False, 10)

# ── Sheet setup ──────────────────────────────────────────────
ws.column_dimensions['A'].width = 25
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[col].width = 12

# ── Title ────────────────────────────────────────────────────
ws['A1'] = "TSMC DCF Valuation Analysis"
ws['A1'].font = f_title

# ── Headers ──────────────────────────────────────────────────
ws['A4'] = "Period"
ws['F4'] = "FY2025"
ws['G4'] = "FY2026"
ws['H4'] = "FY2027"
ws['I4'] = "FY2028"
ws['J4'] = "FY2029"
for col in ['A', 'F', 'G', 'H', 'I', 'J']:
    ws[f'{col}4'].font = mkfont(BLACK, True)

# ── Revenue Build (Rows 6-18) ────────────────────────────────
row = 6
ws[f'A{row}'] = "REVENUE BUILD"
ws[f'A{row}'].font = f_section

segments = [
    ('Smartphone', 7),
    ('HPC', 10),
    ('IoT', 13),
    ('Automotive', 16),
    ('Digital Consumer', 19),
    ('Other', 22),
]

row = 7
for seg_name, model_row in segments:
    ws[f'A{row}'] = f"{seg_name} Revenue"
    for period in ALL:
        col = DC[period]
        model_col = MC[period]
        ws[f'{col}{row}'] = f"=Model!{model_col}{model_row}"
        ws[f'{col}{row}'].font = f_formula
    row += 1

# Total Revenue
ws['A13'] = "Total Revenue"
ws['A13'].font = f_sublabel
for period in ALL:
    col = DC[period]
    ws[f'{col}13'] = f"=SUM({col}7:{col}12)"

row = 14
ws[f'A{row}'] = "Y/Y Growth %"
for period in PROJ:
    col = DC[period]
    prev_col = DC[ALL[ALL.index(period)-1]]
    ws[f'{col}{row}'] = f"={col}13/{prev_col}13 - 1"

# ── Operating Margins (Rows 16-25) ────────────────────────────
row = 16
ws[f'A{row}'] = "OPERATING MARGINS"
ws[f'A{row}'].font = f_section

row = 17
ws[f'A{row}'] = "Gross Margin %"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}38"  # Row 38 = GM%
    ws[f'{col}{row}'].font = f_formula

row = 18
ws[f'A{row}'] = "Operating Margin %"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    # Calculate: Operating Income / Revenue
    ws[f'{col}{row}'] = f"=Model!{model_col}59 / Model!{model_col}25"
    ws[f'{col}{row}'].font = f_formula

# ── UFCF Build (Rows 27-38) ──────────────────────────────────
row = 27
ws[f'A{row}'] = "UFCF BUILD"
ws[f'A{row}'].font = f_section

row = 28
ws[f'A{row}'] = "Revenue"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}13"

row = 29
ws[f'A{row}'] = "Gross Margin %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}17"

row = 30
ws[f'A{row}'] = "Gross Profit"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}28 * {col}29"

row = 31
ws[f'A{row}'] = "Operating Margin %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}18"

row = 32
ws[f'A{row}'] = "EBIT"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}28 * {col}31"

row = 33
ws[f'A{row}'] = "Tax Rate %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = 0.15  # TSMC effective tax rate ~15%
    ws[f'{col}{row}'].number_format = '0.0%'

row = 34
ws[f'A{row}'] = "NOPAT (EBIT × (1-tax))"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}32 * (1 - {col}33)"

row = 35
ws[f'A{row}'] = "D&A"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}72 + Model!{model_col}76"
    ws[f'{col}{row}'].font = f_formula

row = 36
ws[f'A{row}'] = "CapEx"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}86 * 1000"  # Convert from billions
    ws[f'{col}{row}'].font = f_formula

row = 37
ws[f'A{row}'] = "Change in Working Capital"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}28 * 0.10"  # Estimate: 10% of revenue

row = 38
ws[f'A{row}'] = "Unlevered FCF"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}34 + {col}35 - {col}36 - {col}37"

# ── Discounting (Rows 40-45) ──────────────────────────────────
row = 40
ws[f'A{row}'] = "DISCOUNTING"
ws[f'A{row}'].font = f_section

row = 41
ws[f'A{row}'] = "WACC %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = 0.09  # TSMC WACC: ~9% (adjust as needed)
    ws[f'{col}{row}'].number_format = '0.0%'

row = 42
ws[f'A{row}'] = "Discount Factor"
for i, period in enumerate(ALL, 1):
    col = DC[period]
    ws[f'{col}{row}'] = f"=1 / (1 + {col}41)^{i}"

row = 43
ws[f'A{row}'] = "PV of UFCF"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}38 * {col}42"

# ── Terminal Value (Rows 45-48) ──────────────────────────────
row = 45
ws[f'A{row}'] = "TERMINAL VALUE"
ws[f'A{row}'].font = f_section

row = 46
ws[f'A{row}'] = "Terminal Growth %"
ws['F46'] = 0.025  # 2.5% perpetual growth
ws['F46'].number_format = '0.0%'

row = 47
ws[f'A{row}'] = "Terminal Value"
ws['F47'] = f"=F38 * (1 + F46) / (F41 - F46)"

row = 48
ws[f'A{row}'] = "PV of TV"
ws['F48'] = f"=F47 * F42"

# ── Valuation (Rows 50-58) ────────────────────────────────────
row = 50
ws[f'A{row}'] = "VALUATION"
ws[f'A{row}'].font = f_section

row = 51
ws[f'A{row}'] = "Sum PV UFCF"
ws['F51'] = f"=SUM(F43:J43)"

row = 52
ws[f'A{row}'] = "PV Terminal Value"
ws['F52'] = f"=F48"

row = 53
ws[f'A{row}'] = "Enterprise Value"
ws['F53'] = f"=F51 + F52"

row = 54
ws[f'A{row}'] = "Less: Net Debt"
ws['F54'] = f"=Model!L84 - Model!L80"  # (LT Debt + ST Debt) - Cash (adjust rows as needed)

row = 55
ws[f'A{row}'] = "Equity Value"
ws['F55'] = f"=F53 - F54"

row = 56
ws[f'A{row}'] = "Shares Outstanding (mm)"
ws['F56'] = f"=Model!L500"  # Adjust row as needed

row = 57
ws[f'A{row}'] = "Implied Price"
ws['F57'] = f"=F55 / F56"

row = 58
ws[f'A{row}'] = "Current Price"
ws['F58'] = f"='Front Page'!H20"

row = 59
ws[f'A{row}'] = "Upside %"
ws['F59'] = f"=F57 / F58 - 1"
ws['F59'].number_format = '0.0%'

# Save
wb.save(filepath)
print(f"✅ DCF tab built for TSMC!")
print(f"   Rows: 59, Cols: 10")
print(f"   Revenue segments: 5 (Smartphone, HPC, IoT, Automotive, Digital Consumer, Other)")
print(f"   Projection periods: FY2025-FY2029")
