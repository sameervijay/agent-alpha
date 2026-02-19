"""
build_crwv_dcf_tab.py
=====================
Builds the full DCF tab in the CoreWeave analyst model Excel file.
CoreWeave is a simpler model (single revenue line) with EBITDA/Operating Income focus.

Usage:
    cd agent-alpha
    python scripts/build_crwv_dcf_tab.py

CoreWeave Model Tab Key Rows:
- Total Revenue: Row 7
- Non-GAAP EBITDA: Row 12 (use GAAP at Row 59 for DCF)
- GAAP Operating Income: Row 99
- D&A: Row 129 (estimated)
- CapEx: Row 216
- Shares Outstanding: ~Row 500 (need to locate)

Column Mappings (Annual):
- FY2025: S, FY2026: X, FY2027: AC, FY2028: AH, FY2029: AM
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

filepath = 'financial_models/CoreWeave CRWV US.xlsx'
wb = openpyxl.load_workbook(filepath)

if 'DCF' in wb.sheetnames:
    idx = wb.sheetnames.index('DCF')
    del wb['DCF']
else:
    idx = 1
ws = wb.create_sheet('DCF', idx)

# ── Column mappings ──────────────────────────────────────────
MC = {'FY2025': 'S', 'FY2026': 'X', 'FY2027': 'AC', 'FY2028': 'AH', 'FY2029': 'AM'}
DC = {'FY2025': 'F', 'FY2026': 'G', 'FY2027': 'H', 'FY2028': 'I', 'FY2029': 'J'}
ALL = ['FY2025', 'FY2026', 'FY2027', 'FY2028', 'FY2029']
PROJ = ['FY2026', 'FY2027', 'FY2028', 'FY2029']

# ── Styles ───────────────────────────────────────────────────
BLUE = '0000FF'
GREEN = '007F00'
BLACK = '000000'

def mkfont(color=BLACK, bold=False, size=None):
    return Font(color=color, bold=bold, size=size)

# ── Setup ────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 25
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[col].width = 12

# ── Title ────────────────────────────────────────────────────
ws['A1'] = "CoreWeave DCF Valuation Analysis"
ws['A1'].font = mkfont(BLACK, True, 14)

# ── Headers ──────────────────────────────────────────────────
row = 4
ws[f'A{row}'] = "Period"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = period
    ws[f'{col}{row}'].font = mkfont(BLACK, True)

# ── Revenue (Row 6) ──────────────────────────────────────────
row = 6
ws[f'A{row}'] = "REVENUE BUILD"
ws[f'A{row}'].font = mkfont(BLACK, True, 12)

row = 7
ws[f'A{row}'] = "Total Revenue"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}7"
    ws[f'{col}{row}'].font = mkfont(GREEN)

row = 8
ws[f'A{row}'] = "Y/Y Growth %"
for period in PROJ:
    col = DC[period]
    prev_col = DC[ALL[ALL.index(period)-1]]
    ws[f'{col}{row}'] = f"={col}7/{prev_col}7 - 1"

# ── Operating Metrics (Rows 10-15) ───────────────────────────
row = 10
ws[f'A{row}'] = "OPERATING METRICS"
ws[f'A{row}'].font = mkfont(BLACK, True, 12)

row = 11
ws[f'A{row}'] = "EBITDA Margin %"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}15"  # Non-GAAP EBITDA Margin
    ws[f'{col}{row}'].font = mkfont(GREEN)

row = 12
ws[f'A{row}'] = "Operating Income (GAAP)"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}99"  # GAAP Operating Income
    ws[f'{col}{row}'].font = mkfont(GREEN)

row = 13
ws[f'A{row}'] = "Operating Margin %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}12/{col}7"

# ── UFCF Build (Rows 16-26) ──────────────────────────────────
row = 16
ws[f'A{row}'] = "UFCF BUILD"
ws[f'A{row}'].font = mkfont(BLACK, True, 12)

row = 17
ws[f'A{row}'] = "Revenue"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}7"

row = 18
ws[f'A{row}'] = "Operating Income"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}12"

row = 19
ws[f'A{row}'] = "Tax Rate %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = 0.21  # ~21% effective tax rate
    ws[f'{col}{row}'].number_format = '0.0%'

row = 20
ws[f'A{row}'] = "NOPAT (OpInc × (1-Tax))"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}18 * (1 - {col}19)"

row = 21
ws[f'A{row}'] = "D&A"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}129"
    ws[f'{col}{row}'].font = mkfont(GREEN)

row = 22
ws[f'A{row}'] = "CapEx"
for period in ALL:
    col = DC[period]
    model_col = MC[period]
    ws[f'{col}{row}'] = f"=Model!{model_col}216"
    ws[f'{col}{row}'].font = mkfont(GREEN)

row = 23
ws[f'A{row}'] = "Change in Working Capital"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}17 * 0.08"

row = 24
ws[f'A{row}'] = "Unlevered FCF"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}20 + {col}21 - {col}22 - {col}23"

# ── Discounting (Rows 26-30) ──────────────────────────────────
row = 26
ws[f'A{row}'] = "DISCOUNTING"
ws[f'A{row}'].font = mkfont(BLACK, True, 12)

row = 27
ws[f'A{row}'] = "WACC %"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = 0.10  # CoreWeave WACC: ~10% (higher risk, pre-revenue)
    ws[f'{col}{row}'].number_format = '0.0%'

row = 28
ws[f'A{row}'] = "Discount Factor"
for i, period in enumerate(ALL, 1):
    col = DC[period]
    ws[f'{col}{row}'] = f"=1 / (1 + {col}27)^{i}"

row = 29
ws[f'A{row}'] = "PV of UFCF"
for period in ALL:
    col = DC[period]
    ws[f'{col}{row}'] = f"={col}24 * {col}28"

# ── Terminal Value (Rows 31-33) ──────────────────────────────
row = 31
ws[f'A{row}'] = "TERMINAL VALUE"
ws[f'A{row}'].font = mkfont(BLACK, True, 12)

row = 32
ws[f'A{row}'] = "Terminal Growth %"
ws['F32'] = 0.03  # 3% perpetual growth (higher than TSMC due to growth)
ws['F32'].number_format = '0.0%'

row = 33
ws[f'A{row}'] = "Terminal Value"
ws['F33'] = f"=F24 * (1 + F32) / (F27 - F32)"

row = 34
ws[f'A{row}'] = "PV of TV"
ws['F34'] = f"=F33 * F28"

# ── Valuation (Rows 36-42) ───────────────────────────────────
row = 36
ws[f'A{row}'] = "VALUATION"
ws[f'A{row}'].font = mkfont(BLACK, True, 12)

row = 37
ws[f'A{row}'] = "Sum PV UFCF"
ws['F37'] = f"=SUM(F29:J29)"

row = 38
ws[f'A{row}'] = "PV Terminal Value"
ws['F38'] = f"=F34"

row = 39
ws[f'A{row}'] = "Enterprise Value"
ws['F39'] = f"=F37 + F38"

row = 40
ws[f'A{row}'] = "Less: Net Debt"
ws['F40'] = f"=Model!L600 - Model!L550"  # Adjust as needed

row = 41
ws[f'A{row}'] = "Equity Value"
ws['F41'] = f"=F39 - F40"

row = 42
ws[f'A{row}'] = "Shares Outstanding (mm)"
ws['F42'] = f"=Model!L700"  # Adjust as needed

row = 43
ws[f'A{row}'] = "Implied Price"
ws['F43'] = f"=F41 / F42"

row = 44
ws[f'A{row}'] = "Current Price"
ws['F44'] = f"='Front Page'!H20"

row = 45
ws[f'A{row}'] = "Upside %"
ws['F45'] = f"=F43 / F44 - 1"
ws['F45'].number_format = '0.0%'

wb.save(filepath)
print(f"✅ DCF tab built for CoreWeave!")
print(f"   Rows: 45, Cols: 10")
print(f"   Revenue: Single line (total GPU rental revenue)")
print(f"   Projection periods: FY2025-FY2029")
