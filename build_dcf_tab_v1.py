"""
build_dcf_tab_v1.py
===================
INITIAL version of the DCF tab builder (simpler layout without segment revenue build).
Kept for reference. The current version is build_dcf_tab.py.

This version has:
- Single-line revenue/EBIT from Model tab
- UFCF build, discounting, equity bridge, NWC detail
- No segment breakdown or margin driver rows
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('NVIDIA NVDA US.xlsx')

# Remove existing empty DCF sheet and recreate
if 'DCF' in wb.sheetnames:
    idx = wb.sheetnames.index('DCF')
    del wb['DCF']
else:
    idx = 1

ws = wb.create_sheet('DCF', idx)

# ── Column mappings ──────────────────────────────────────────
# Model tab columns for each fiscal year
MC = {
    'FY2025': 'BS', 'FY2026': 'BX', 'FY2027': 'CC',
    'FY2028': 'CD', 'FY2029': 'CE', 'FY2030': 'CF',
}
# DCF tab columns
DC = {
    'FY2025': 'F', 'FY2026': 'G', 'FY2027': 'H',
    'FY2028': 'I', 'FY2029': 'J', 'FY2030': 'K',
}
ALL_FY = ['FY2025','FY2026','FY2027','FY2028','FY2029','FY2030']
PROJ_FY = ['FY2026','FY2027','FY2028','FY2029','FY2030']

# Model tab row numbers
MR = dict(
    revenue=362, ebit=386, ebitda=392, da=933, capex=934,
    ar=954, inv=958, ap=985, accrued=987,
    cash=952, mkt_sec=953, lt_debt=992, st_debt=988,
    shares_wad=429, sbc=408,
)

# ── Formatting constants ─────────────────────────────────────
bold = Font(bold=True)
title_f = Font(bold=True, size=14)
sub_f = Font(bold=True, size=11)
hdr_f = Font(bold=True, size=10)
ital9 = Font(italic=True, size=9)
note_f = Font(italic=True, size=9, color='666666')
blue_bold = Font(bold=True, color='0000CC', size=11)
pct = '0.0%'
num = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
dlr = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
dec3 = '_(* #,##0.000_);_(* \\(#,##0.000\\);_(* "-"??_);_(@_)'
price = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
multx = '#,##0.0"x"'
ctr = Alignment(horizontal='center')

thin = Side(style='thin')
bottom_border = Border(bottom=thin)
top_bottom = Border(top=thin, bottom=Side(style='double'))

# ── Column widths ────────────────────────────────────────────
widths = {'A':2,'B':42,'C':5,'D':2,'E':2,'F':16,'G':16,'H':16,
          'I':16,'J':16,'K':16,'L':3,'M':30,'N':2,'O':2,'P':12,
          'Q':2,'R':38}
for c,w in widths.items():
    ws.column_dimensions[c].width = w

# ── Helper to set a cell ─────────────────────────────────────
def cell(addr, val, font=None, fmt=None, align=None, border=None):
    ws[addr] = val
    if font:  ws[addr].font = font
    if fmt:   ws[addr].number_format = fmt
    if align: ws[addr].alignment = align
    if border: ws[addr].border = border

# ══════════════════════════════════════════════════════════════
# TITLE
# ══════════════════════════════════════════════════════════════
cell('B1', 'NVIDIA Corporation', title_f)
cell('B2', 'Discounted Cash Flow Valuation Analysis', sub_f)

# ── Section headers ──────────────────────────────────────────
cell('F5', 'Historical', hdr_f, align=ctr)
cell('G5', 'Projected', hdr_f, align=ctr)
cell('M5', 'Assumptions', Font(bold=True, size=12, underline='single'))
cell('R5', 'Notes', Font(bold=True, size=12, underline='single'))

cell('B7', '($ in millions)', ital9)
for fy in ALL_FY:
    cell(f'{DC[fy]}7', fy, bold, align=ctr, border=bottom_border)

# ══════════════════════════════════════════════════════════════
# ASSUMPTIONS PANEL  (M6:R16)
# ══════════════════════════════════════════════════════════════
assumptions = [
    (6,  'Risk Free Rate:',             0.0425, '0.00%', '10-Year US Treasury yield'),
    (7,  'Market Risk Premium:',        0.055,  '0.00%', 'Damodaran equity risk premium'),
    (8,  'Unlevered Beta:',             1.65,   '0.00',  'Semiconductor industry avg'),
    (9,  'Projected Equity to Value:',  0.998,  '0.0%',  'NVDA ~100% equity-financed'),
    (10, 'Levered Beta:',              '=(1+(1-P12)*(1-P9)/P9)*P8', '0.00', 'Hamada equation'),
    (11, 'Cost of Equity:',            '=P6+P7*P10', '0.00%', 'CAPM: Rf + B(Rm-Rf)'),
    (12, 'Tax Rate:',                   0.15,   '0%',    'Long-term effective tax rate'),
    (13, 'Cost of Debt:',              0.035,  '0.00%', 'Based on outstanding senior notes'),
    (14, 'WACC:',                      '=P11*P9+P13*(1-P12)*(1-P9)', '0.0%',
          'Rd*(1-T)*(D/V) + Re*(E/V)'),
    (15, 'Terminal Growth Rate:',       0.03,   '0.00%', 'Long-term nominal GDP growth'),
    (16, 'Implied TV/EBITDA:',         f"=F25/Model!{MC['FY2030']}{MR['ebitda']}",
          multx, 'Terminal value / FY2030 EBITDA'),
]
for r, label, val, fmt, note in assumptions:
    cell(f'M{r}', label, bold)
    cell(f'P{r}', val, fmt=fmt)
    cell(f'R{r}', note, note_f)

# ══════════════════════════════════════════════════════════════
# UFCF BUILD  (Rows 9-19)
# ══════════════════════════════════════════════════════════════

# Row 9: Revenue
cell('B9', 'Revenue', bold)
for fy in ALL_FY:
    cell(f'{DC[fy]}9', f"=Model!{MC[fy]}{MR['revenue']}", fmt=num)

# Row 10: Y/Y Revenue Growth
cell('B10', 'Y/Y Revenue Growth')
for fy in PROJ_FY:
    d = DC[fy]; p = DC[ALL_FY[ALL_FY.index(fy)-1]]
    cell(f'{d}10', f'={d}9/{p}9-1', fmt=pct)

# Row 12: Non-GAAP EBIT
cell('B12', 'Non-GAAP Operating Income (EBIT)', bold)
for fy in ALL_FY:
    cell(f'{DC[fy]}12', f"=Model!{MC[fy]}{MR['ebit']}", fmt=num)

# Row 13: EBIT Margin
cell('B13', 'Non-GAAP EBIT Margin')
for fy in ALL_FY:
    d = DC[fy]
    cell(f'{d}13', f'={d}12/{d}9', fmt=pct)

# Row 14: Less Taxes on EBIT  (projected only)
cell('B14', 'Less: Taxes on EBIT')
for fy in PROJ_FY:
    cell(f'{DC[fy]}14', f'=-{DC[fy]}12*$P$12', fmt=num)

# Row 15: NOPAT
cell('B15', 'NOPAT', bold)
for fy in PROJ_FY:
    d = DC[fy]
    cell(f'{d}15', f'={d}12+{d}14', fmt=num)

# Row 16: Plus D&A
cell('B16', 'Plus: Depreciation & Amortization')
for fy in PROJ_FY:
    cell(f'{DC[fy]}16', f"=Model!{MC[fy]}{MR['da']}", fmt=num)

# Row 17: Less CapEx  (sign-flipped)
cell('B17', 'Less: Capital Expenditures')
for fy in PROJ_FY:
    cell(f'{DC[fy]}17', f"=-Model!{MC[fy]}{MR['capex']}", fmt=num)

# Row 18: Less Increase in NWC  (references row 51 below)
cell('B18', 'Less: Increase in Net Working Capital')
for fy in PROJ_FY:
    cell(f'{DC[fy]}18', f'=-{DC[fy]}51', fmt=num)

# Row 19: UFCF  (sum of 15:18)
cell('B19', 'Unlevered Free Cash Flow', bold, border=top_bottom)
for fy in PROJ_FY:
    d = DC[fy]
    cell(f'{d}19', f'=SUM({d}15:{d}18)', fmt=num, border=top_bottom)

# ══════════════════════════════════════════════════════════════
# DISCOUNTING  (Rows 21-28)
# ══════════════════════════════════════════════════════════════

# Row 21: Discount Factor
cell('B21', 'Times: Discount Factor')
for i, fy in enumerate(PROJ_FY, 1):
    cell(f'{DC[fy]}21', f'=1/(1+$P$14)^{i}', fmt=dec3)

# Row 22: Discounted CFs
cell('B22', 'Discounted Cash Flows')
for fy in PROJ_FY:
    d = DC[fy]
    cell(f'{d}22', f'={d}19*{d}21', fmt=num)

# Row 24: Sum of Discounted CFs
cell('B24', 'Sum of Discounted Cash Flows', bold)
cell('F24', '=SUM(G22:K22)', fmt=dlr)

# Row 25: Terminal Value  (Gordon Growth)
cell('B25', 'Terminal Value:', bold)
cell('F25', '=(K19*(1+$P$15))/($P$14-$P$15)', fmt=num)

# Row 26: Times Discount Factor (= year 5 DF)
cell('B26', 'Times: Discount Factor')
cell('F26', '=K21', fmt=dec3)

# Row 27: PV of Terminal Value
cell('B27', 'PV of Terminal Value')
cell('F27', '=F25*F26', fmt=dlr)

# Row 28: Enterprise Value
cell('C28', 'Enterprise Value', Font(bold=True, size=12))
cell('F28', '=F24+F27', fmt=dlr, border=top_bottom)

# ══════════════════════════════════════════════════════════════
# EQUITY BRIDGE  (Rows 30-40)
# ══════════════════════════════════════════════════════════════
cell('B30', 'Equity Bridge:', Font(bold=True, size=11, underline='single'))

cell('B31', 'Enterprise Value')
cell('F31', '=F28', fmt=dlr)

# Total Debt = LT + ST from most recent balance sheet (FY2025)
cell('B32', 'Less: Total Debt')
cell('F32', f"=-(Model!{MC['FY2025']}{MR['lt_debt']}+IF(ISNUMBER(Model!{MC['FY2025']}{MR['st_debt']}),Model!{MC['FY2025']}{MR['st_debt']},0))", fmt=dlr)

cell('B33', 'Plus: Cash & Cash Equivalents')
cell('F33', f"=Model!{MC['FY2025']}{MR['cash']}", fmt=dlr)

cell('B34', 'Plus: Marketable Securities')
cell('F34', f"=Model!{MC['FY2025']}{MR['mkt_sec']}", fmt=dlr)

cell('C35', 'Equity Value', Font(bold=True, size=12))
cell('F35', '=SUM(F31:F34)', fmt=dlr, border=top_bottom)

cell('B36', 'Diluted Shares Outstanding (mm)')
cell('F36', f"=Model!{MC['FY2025']}{MR['shares_wad']}", fmt=num)

cell('C37', 'Implied Share Price ($/share)', Font(bold=True, size=12))
cell('F37', '=F35/F36', fmt=price, border=top_bottom)

# Current Price and Upside/Downside
cell('B39', 'Current Stock Price ($/share)', bold)
cell('F39', "='Front Page'!H20", fmt=price)

cell('B40', 'Upside / (Downside)', blue_bold)
cell('F40', '=F37/F39-1', fmt=pct, font=blue_bold)

# ══════════════════════════════════════════════════════════════
# NWC DETAIL  (Rows 43-53)
# ══════════════════════════════════════════════════════════════
cell('B43', 'Net Working Capital Calculation:', Font(bold=True, size=11, underline='single'))

# Year headers
for fy in ALL_FY:
    cell(f'{DC[fy]}44', fy, bold, align=ctr, border=bottom_border)

# AR
cell('B45', 'Accounts Receivable')
for fy in ALL_FY:
    cell(f'{DC[fy]}45', f"=Model!{MC[fy]}{MR['ar']}", fmt=num)

# Inventory
cell('B46', 'Inventories')
for fy in ALL_FY:
    cell(f'{DC[fy]}46', f"=Model!{MC[fy]}{MR['inv']}", fmt=num)

# AP (negative)
cell('B47', 'Less: Accounts Payable')
for fy in ALL_FY:
    cell(f'{DC[fy]}47', f"=-Model!{MC[fy]}{MR['ap']}", fmt=num)

# Accrued (negative)
cell('B48', 'Less: Accrued & Other Current Liabilities')
for fy in ALL_FY:
    cell(f'{DC[fy]}48', f"=-Model!{MC[fy]}{MR['accrued']}", fmt=num)

# NWC total
cell('B49', 'Net Working Capital', bold, border=bottom_border)
for fy in ALL_FY:
    d = DC[fy]
    cell(f'{d}49', f'=SUM({d}45:{d}48)', fmt=num, border=bottom_border)

# Change in NWC (proj only)
cell('B51', 'Change in Net Working Capital', bold)
for fy in PROJ_FY:
    d = DC[fy]; p = DC[ALL_FY[ALL_FY.index(fy)-1]]
    cell(f'{d}51', f'={d}49-{p}49', fmt=num)

# UFCF Growth
cell('B53', 'UFCF Growth Rate')
for i, fy in enumerate(PROJ_FY):
    if i > 0:
        d = DC[fy]; p = DC[PROJ_FY[i-1]]
        cell(f'{d}53', f'={d}19/{p}19-1', fmt=pct)

# ══════════════════════════════════════════════════════════════
# MEMO ITEMS  (Row 55-57)
# ══════════════════════════════════════════════════════════════
cell('B55', 'Memo:', Font(bold=True, size=11, underline='single'))
cell('B56', 'Stock-Based Compensation (excluded from UFCF)')
for fy in PROJ_FY:
    cell(f'{DC[fy]}56', f"=Model!{MC[fy]}{MR['sbc']}", fmt=num)

cell('B57', 'Non-GAAP EBITDA')
for fy in ALL_FY:
    cell(f'{DC[fy]}57', f"=Model!{MC[fy]}{MR['ebitda']}", fmt=num)

# ── Save ─────────────────────────────────────────────────────
wb.save('NVIDIA NVDA US.xlsx')
print('DCF tab created successfully!')
print('Verify by opening the file in Excel.')
