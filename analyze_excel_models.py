#!/usr/bin/env python3
"""
Analyze Excel financial models to identify key DCF cells.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def get_column_letter_from_index(col_index):
    """Convert 1-based column index to letter."""
    return get_column_letter(col_index)

def analyze_excel_file(filepath):
    """Analyze an Excel financial model and extract key cells."""
    print(f"\n{'='*100}")
    print(f"Analyzing: {os.path.basename(filepath)}")
    print(f"{'='*100}\n")

    try:
        # Load workbook
        wb = openpyxl.load_workbook(filepath, data_only=False)

        # List all sheets
        print(f"Available sheets: {wb.sheetnames}\n")

        # Try to find and analyze the Model sheet
        if 'Model' in wb.sheetnames:
            ws = wb['Model']
            print(f"Analyzing 'Model' sheet: {ws.dimensions}\n")

            print("=" * 100)
            print("MODEL STRUCTURE (Rows 1-150, showing content)")
            print("=" * 100)

            for row_idx in range(1, min(151, ws.max_row + 1)):
                row_data = []
                has_content = False

                for col_idx in range(1, min(35, ws.max_column + 1)):  # First 35 columns
                    cell = ws.cell(row_idx, col_idx)
                    value = cell.value
                    col_letter = get_column_letter(col_idx)

                    if value is not None:
                        has_content = True
                        # Truncate long values
                        val_str = str(value)[:40] if value else ""
                        if len(val_str) > 40:
                            val_str = val_str[:37] + "..."
                        row_data.append(f"{col_letter}: {val_str}")

                if has_content:
                    print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

        # Also check other sheets for DCF or Valuation
        for sheet_name in wb.sheetnames:
            if 'DCF' in sheet_name or 'Valuation' in sheet_name or 'Summary' in sheet_name:
                ws = wb[sheet_name]
                print(f"\n{'='*100}")
                print(f"Found '{sheet_name}' sheet - showing first 80 rows")
                print(f"{'='*100}\n")

                for row_idx in range(1, min(81, ws.max_row + 1)):
                    row_data = []
                    has_content = False

                    for col_idx in range(1, min(35, ws.max_column + 1)):
                        cell = ws.cell(row_idx, col_idx)
                        value = cell.value
                        col_letter = get_column_letter(col_idx)

                        if value is not None:
                            has_content = True
                            val_str = str(value)[:40] if value else ""
                            if len(val_str) > 40:
                                val_str = val_str[:37] + "..."
                            row_data.append(f"{col_letter}: {val_str}")

                    if has_content:
                        print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

        wb.close()

    except Exception as e:
        print(f"ERROR analyzing file: {e}")
        import traceback
        traceback.print_exc()

# Analyze the three target companies
companies = {
    'TSMC': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/Taiwan Semiconductor Manufacturing Company TSM US.xlsx',
    'CoreWeave': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/CoreWeave CRWV US.xlsx',
    'ASML': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/ASML Holding ASML NA.xlsx',
}

for company, filepath in companies.items():
    if os.path.exists(filepath):
        analyze_excel_file(filepath)
    else:
        print(f"File not found: {filepath}")
