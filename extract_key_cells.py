#!/usr/bin/env python3
"""
Extract specific key cell references for DCF analysis.
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os

def get_value(ws, row, col_letter):
    """Get cell value safely."""
    try:
        return ws[f"{col_letter}{row}"].value
    except:
        return None

def analyze_company_cells(filepath, company_name):
    """Extract key cells for DCF analysis."""
    print(f"\n{'='*120}")
    print(f"DCF CELL MAPPING: {company_name}")
    print(f"File: {os.path.basename(filepath)}")
    print(f"{'='*120}\n")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)

        if 'Model' not in wb.sheetnames:
            print("ERROR: No 'Model' sheet found")
            return

        ws = wb['Model']

        # Get column header row (typically row 5)
        print("COLUMN STRUCTURE")
        print("-" * 120)
        header_row = 5
        headers = {}
        for col_idx in range(1, min(ws.max_column + 1, 90)):
            cell_val = ws.cell(header_row, col_idx).value
            if cell_val:
                col_letter = get_column_letter(col_idx)
                headers[col_letter] = cell_val
                # Only print first 30 columns
                if col_idx <= 30:
                    print(f"  {col_letter}: {str(cell_val)[:50]}")

        print(f"\nTotal columns with headers: {len(headers)}")

        # Find key metric rows
        print("\n" + "=" * 120)
        print("KEY METRICS LOCATIONS")
        print("=" * 120)

        key_metrics = {
            'Revenue': [],
            'Gross Profit/Margin': [],
            'Operating Expenses (R&D, SG&A, etc)': [],
            'Operating Income (EBIT/EBITDA)': [],
            'Net Income': [],
            'Depreciation & Amortization': [],
            'CapEx': [],
            'Working Capital': [],
            'Cash & Debt': [],
            'Shares Outstanding': [],
            'Tax': []
        }

        for row_idx in range(1, min(150, ws.max_row + 1)):
            cell_a = ws.cell(row_idx, 1).value
            if not cell_a:
                continue

            cell_text = str(cell_a).lower()

            # Categorize metrics
            if 'revenue' in cell_text or 'total' in cell_text and 'revenue' in cell_text:
                key_metrics['Revenue'].append((row_idx, cell_a))
            elif 'gross' in cell_text and ('profit' in cell_text or 'margin' in cell_text):
                key_metrics['Gross Profit/Margin'].append((row_idx, cell_a))
            elif any(x in cell_text for x in ['r&d', 'research', 'sga', 'selling', 'administrative', 'opex']):
                key_metrics['Operating Expenses (R&D, SG&A, etc)'].append((row_idx, cell_a))
            elif any(x in cell_text for x in ['ebit', 'operating income', 'ebitda']):
                key_metrics['Operating Income (EBIT/EBITDA)'].append((row_idx, cell_a))
            elif 'net income' in cell_text or 'net loss' in cell_text:
                key_metrics['Net Income'].append((row_idx, cell_a))
            elif any(x in cell_text for x in ['depreciation', 'amortization', 'd&a']):
                key_metrics['Depreciation & Amortization'].append((row_idx, cell_a))
            elif 'capex' in cell_text or 'capital expenditure' in cell_text:
                key_metrics['CapEx'].append((row_idx, cell_a))
            elif any(x in cell_text for x in ['working capital', 'nwc', 'accounts receivable', 'inventory', 'payable']):
                key_metrics['Working Capital'].append((row_idx, cell_a))
            elif any(x in cell_text for x in ['cash', 'debt', 'borrowing']):
                key_metrics['Cash & Debt'].append((row_idx, cell_a))
            elif any(x in cell_text for x in ['share', 'diluted', 'wad']):
                key_metrics['Shares Outstanding'].append((row_idx, cell_a))
            elif 'tax' in cell_text or 'provision' in cell_text:
                key_metrics['Tax'].append((row_idx, cell_a))

        # Print key metrics organized by category
        for category, metrics in key_metrics.items():
            if metrics:
                print(f"\n{category}:")
                for row_num, metric_name in metrics[:10]:  # Limit to first 10 per category
                    print(f"  Row {row_num}: {str(metric_name)[:80]}")

        wb.close()

    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

# Analyze companies
companies = {
    'TSMC': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/Taiwan Semiconductor Manufacturing Company TSM US.xlsx',
    'CoreWeave': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/CoreWeave CRWV US.xlsx',
    'ASML': '/Users/matthewwolfman/Documents/CS372_Assignment3/agent-alpha/financial_models/ASML Holding ASML NA.xlsx',
}

for company, filepath in companies.items():
    if os.path.exists(filepath):
        analyze_company_cells(filepath, company)
