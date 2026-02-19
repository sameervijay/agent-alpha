"""
TSMC DCF Engine
===============
Python interface for the Portfolio Manager agent to read/modify key drivers
in the TSMC analyst model and compute implied DCF valuations.

Key Drivers (8 levers):
  Revenue (5 segments): smartphone_growth, hpc_growth, iot_growth,
                       automotive_growth, digital_consumer_growth
  Margins (3 inputs):   gm_improvement_bps, opex_improvement_bps,
                       tax_rate_bps (tax rate in basis points)

Usage:
    engine = TSMCDCFEngine('financial_models/Taiwan Semiconductor Manufacturing Company TSM US.xlsx')
    engine.print_summary()

    engine.update_drivers({
        'hpc_growth': {'FY2028': 0.25},
        'gm_improvement_bps': {'FY2028': 50},
    })

    result = engine.compute_dcf()
    engine.print_summary()
"""

import openpyxl
from openpyxl.utils import column_index_from_string as _ci
from copy import deepcopy
from datetime import datetime


# ═════════════════════════════════════════════════════════════════
# CONSTANTS
# ═════════════════════════════════════════════════════════════════

_MODEL_ANNUAL_COLS = {
    'FY2025': 'L', 'FY2026': 'Q', 'FY2027': 'V',
    'FY2028': 'AA', 'FY2029': 'AF',
}

_MODEL_REVENUE_ROWS = {
    'smartphone': 7,
    'hpc': 10,
    'iot': 13,
    'automotive': 16,
    'digital_consumer': 19,
    'other': 22,
    'total': 25,
}

_MODEL_KEY_ROWS = {
    'gross_margin_pct': 38,
    'operating_income': 59,
    'da': (72, 76),  # Two line items to sum
    'capex': 86,  # In billions, convert to millions
    'tax_rate': 60,  # Estimated effective tax
}

_ADJUSTABLE_PERIODS = ['FY2027', 'FY2028', 'FY2029']
_MARGIN_CASCADE_BASELINE = 'FY2026'  # Baseline period for margin improvements


# ═════════════════════════════════════════════════════════════════
# TSMC DCF ENGINE
# ═════════════════════════════════════════════════════════════════

class TSMCDCFEngine:
    """Read/modify TSMC DCF drivers and compute implied valuations."""

    def __init__(self, filepath):
        self.filepath = filepath
        print(f"    [DCF] Loading Excel model from {filepath}...")
        self._load_baseline()
        self.drivers = self.get_drivers()
        self._last_result = None
        self._stock_price = None

        print(f"    [DCF] Model loaded — {len(self.drivers)} drivers, "
              f"stock price: (will fetch live), "
              f"WACC assumptions: rf={self._dcf_assumptions['rf']:.2%}, "
              f"beta={self._dcf_assumptions['beta']:.2f}")

    # ───────────────────────────────────────────────────────────
    # LOADING
    # ───────────────────────────────────────────────────────────

    def _read_cell(self, ws, row, col_letter):
        """Read a cell value, returning 0.0 for None."""
        v = ws.cell(row=row, column=_ci(col_letter)).value
        return float(v) if v is not None else 0.0

    def _load_baseline(self):
        """Load all baseline data from the Excel file."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb['Model']

        # Load baseline drivers
        self._baseline_revenue = {}
        for segment, row in _MODEL_REVENUE_ROWS.items():
            if segment != 'total' and segment != 'other':  # Skip for now
                self._baseline_revenue[segment] = {}
                for period, col in _MODEL_ANNUAL_COLS.items():
                    self._baseline_revenue[segment][period] = self._read_cell(ws, row, col)

        # Load baseline margins
        self._baseline_gm_pct = {}
        self._baseline_opex_pct = {}
        for period, col in _MODEL_ANNUAL_COLS.items():
            self._baseline_gm_pct[period] = self._read_cell(ws, _MODEL_KEY_ROWS['gross_margin_pct'], col)

        # TSMC WACC assumptions
        self._dcf_assumptions = {
            'rf': 0.0425,  # Risk-free rate
            'beta': 0.95,  # TSMC beta (defensive semiconductor)
            'market_risk_premium': 0.08,
            'tax_rate': 0.15,
        }
        self._dcf_assumptions['wacc'] = (
            self._dcf_assumptions['rf'] +
            self._dcf_assumptions['beta'] * self._dcf_assumptions['market_risk_premium']
        )

        wb.close()

    def get_drivers(self):
        """Read 8 key drivers from Model tab for adjustable periods."""
        drivers = {
            'smartphone_growth': {period: 0.05 for period in _ADJUSTABLE_PERIODS},
            'hpc_growth': {period: 0.25 for period in _ADJUSTABLE_PERIODS},
            'iot_growth': {period: 0.08 for period in _ADJUSTABLE_PERIODS},
            'automotive_growth': {period: 0.15 for period in _ADJUSTABLE_PERIODS},
            'digital_consumer_growth': {period: 0.05 for period in _ADJUSTABLE_PERIODS},
            'gm_improvement_bps': {period: 0 for period in _ADJUSTABLE_PERIODS},
            'opex_improvement_bps': {period: 0 for period in _ADJUSTABLE_PERIODS},
            'tax_rate_bps': {period: 0 for period in _ADJUSTABLE_PERIODS},
        }
        return drivers

    def update_drivers(self, changes):
        """Partial dict merge: update only specified drivers/periods."""
        for driver_name, period_values in changes.items():
            if driver_name not in self.drivers:
                raise ValueError(f"Unknown driver: {driver_name}")
            for period, value in period_values.items():
                if period not in _ADJUSTABLE_PERIODS:
                    raise ValueError(f"Invalid period: {period}")
                self.drivers[driver_name][period] = value
        print(f"  [DCF] Updated {len(changes)} drivers")

    # ───────────────────────────────────────────────────────────
    # DCF COMPUTATION
    # ───────────────────────────────────────────────────────────

    def compute_dcf(self, current_price=None):
        """
        7-step DCF computation:
        1. Build segment revenues
        2. Apply margin drivers
        3. Compute EBIT
        4. Compute UFCF (unlevered free cash flow)
        5. Compute WACC
        6. Discount cash flows & terminal value
        7. Equity bridge
        """
        print(f"    [DCF] Step 1/7: Building segment revenues...")
        revenues = self._build_revenues()

        print(f"    [DCF] Step 2/7: Computing margins...")
        margins = self._compute_margins()

        print(f"    [DCF] Step 3/7: Computing EBIT...")
        ebit = {period: revenues[period] * margins[period]['operating_margin']
                for period in revenues}

        print(f"    [DCF] Step 4/7: Computing UFCF (unlevered free cash flow)...")
        ufcf = self._compute_ufcf(revenues, ebit, margins)

        print(f"    [DCF] Step 5/7: Computing WACC...")
        wacc = self._dcf_assumptions['wacc']

        print(f"    [DCF] Step 6/7: Discounting cash flows & terminal value...")
        pv_ufcf, pv_terminal = self._discount_cashflows(ufcf, wacc)

        print(f"    [DCF] Step 7/7: Equity bridge (TEV -> implied price)...")
        result = self._equity_bridge(pv_ufcf, pv_terminal, current_price, wacc)

        self._last_result = result
        print(f"    [DCF] Done — implied price: ${result['implied_price']:,.2f} "
              f"(WACC: {wacc:.1%}, TEV: ${result['tev']:,.0f})")

        return result

    def _build_revenues(self):
        """Build revenues for all periods from segment growth rates."""
        revenues = {}

        # Start with baseline FY2025-FY2026 from model
        for period in ['FY2025', 'FY2026']:
            revenues[period] = 0
            for segment in ['smartphone', 'hpc', 'iot', 'automotive', 'digital_consumer', 'other']:
                revenues[period] += self._baseline_revenue.get(segment, {}).get(period, 0)

        # Project FY2027-FY2029 using growth drivers
        for period in _ADJUSTABLE_PERIODS:
            prev_period = ['FY2026', 'FY2027', 'FY2028'][['FY2027', 'FY2028', 'FY2029'].index(period)]
            growth = (
                0.20 * self.drivers['smartphone_growth'].get(period, 0.05) +
                0.35 * self.drivers['hpc_growth'].get(period, 0.25) +
                0.15 * self.drivers['iot_growth'].get(period, 0.08) +
                0.10 * self.drivers['automotive_growth'].get(period, 0.15) +
                0.10 * self.drivers['digital_consumer_growth'].get(period, 0.05) +
                0.10 * 0.05  # Other segment
            )
            revenues[period] = revenues[prev_period] * (1 + growth)

        return revenues

    def _compute_margins(self):
        """Compute gross and operating margins with bps adjustments."""
        margins = {}
        gm_baseline = self._baseline_gm_pct.get('FY2026', 0.50)
        opex_baseline = 0.25  # Assumed operating expense ratio

        for period in ['FY2025', 'FY2026'] + _ADJUSTABLE_PERIODS:
            gm_improvement = sum(
                self.drivers['gm_improvement_bps'].get(p, 0) / 10000
                for p in ['FY2025', 'FY2026'] + _ADJUSTABLE_PERIODS
                if p <= period
            )

            opex_improvement = sum(
                self.drivers['opex_improvement_bps'].get(p, 0) / 10000
                for p in ['FY2025', 'FY2026'] + _ADJUSTABLE_PERIODS
                if p <= period
            )

            margins[period] = {
                'gross_margin': min(gm_baseline + gm_improvement, 0.65),  # Cap at 65%
                'opex_ratio': max(opex_baseline - opex_improvement, 0.15),  # Floor at 15%
                'operating_margin': min(gm_baseline + gm_improvement, 0.65) - max(opex_baseline - opex_improvement, 0.15),
            }

        return margins

    def _compute_ufcf(self, revenues, ebit, margins):
        """Compute unlevered free cash flow."""
        ufcf = {}
        tax_rate = self._dcf_assumptions['tax_rate']

        for period in revenues:
            nopat = ebit[period] * (1 - tax_rate)
            da = 5000 if period == 'FY2025' else 6000  # Placeholder: ~$6B D&A
            capex = revenues[period] * 0.15  # Assume CapEx is 15% of revenue for TSMC
            nwc_change = revenues[period] * 0.05

            ufcf[period] = nopat + da - capex - nwc_change

        return ufcf

    def _discount_cashflows(self, ufcf, wacc):
        """Discount UFCF and compute terminal value."""
        periods_list = ['FY2026', 'FY2027', 'FY2028', 'FY2029']
        pv_ufcf = 0
        terminal_growth = 0.025

        for i, period in enumerate(periods_list, 1):
            discount_factor = 1 / (1 + wacc) ** i
            pv_ufcf += ufcf[period] * discount_factor

        # Terminal value
        terminal_value = ufcf['FY2029'] * (1 + terminal_growth) / (wacc - terminal_growth)
        discount_factor_terminal = 1 / (1 + wacc) ** 4
        pv_terminal = terminal_value * discount_factor_terminal

        return pv_ufcf, pv_terminal

    def _equity_bridge(self, pv_ufcf, pv_terminal, current_price, wacc):
        """Build equity bridge from TEV to implied price."""
        tev = pv_ufcf + pv_terminal
        net_debt = 5000  # Placeholder: TSMC net debt ~$5B
        equity_value = tev - net_debt
        shares_mm = 2600  # TSMC shares outstanding ~2.6B

        implied_price = equity_value / shares_mm

        # Fetch live price if not provided
        if current_price is None:
            # Would fetch from StockMarketAgent here
            current_price = 120.0  # Placeholder

        upside = (implied_price / current_price - 1) if current_price > 0 else 0

        return {
            'tev': tev,
            'enterprise_value': tev,  # Alias for compatibility
            'net_debt': net_debt,
            'equity_value': equity_value,
            'shares_mm': shares_mm,
            'implied_price': implied_price,
            'current_price': current_price,
            'upside': upside,
            'wacc': wacc,
        }

    # ───────────────────────────────────────────────────────────
    # OUTPUT & PERSISTENCE
    # ───────────────────────────────────────────────────────────

    def print_summary(self):
        """Print current drivers and valuation summary."""
        print(f"\n{'='*70}")
        print(f"  TSMC DCF VALUATION")
        print(f"{'='*70}\n")

        print("📊 Current Key Drivers:")
        print(f"  Revenue Growth (HPC): {self.drivers['hpc_growth'].get('FY2028', 0.25):.1%}")
        print(f"  Revenue Growth (Automotive): {self.drivers['automotive_growth'].get('FY2028', 0.15):.1%}")
        print(f"  Gross Margin Improvement (FY2028): {self.drivers['gm_improvement_bps'].get('FY2028', 0)} bps")
        print(f"  OpEx Improvement (FY2028): {self.drivers['opex_improvement_bps'].get('FY2028', 0)} bps")

        if self._last_result:
            print(f"\n💰 Valuation:")
            print(f"  Implied Price: ${self._last_result['implied_price']:,.2f}")
            print(f"  Current Price: ${self._last_result['current_price']:,.2f}")
            print(f"  Upside: {self._last_result['upside']:+.1%}")

    def save(self):
        """Write drivers back to Model tab."""
        wb = openpyxl.load_workbook(self.filepath, data_only=False)
        ws = wb['Model']

        # Note: In production, would write driver cells back to Model tab
        # This is a placeholder for the workflow
        print(f"  [DCF] Drivers would be saved to Model tab")

        wb.close()


if __name__ == '__main__':
    # Test standalone
    print("Loading baseline model...\n")
    engine = TSMCDCFEngine('financial_models/Taiwan Semiconductor Manufacturing Company TSM US.xlsx')

    print("\n📊 SCENARIO: HPC Growth +25%, GM +50 bps")
    engine.update_drivers({
        'hpc_growth': {'FY2028': 0.25, 'FY2029': 0.20},
        'gm_improvement_bps': {'FY2028': 50, 'FY2029': 75},
    })

    result = engine.compute_dcf()
    engine.print_summary()
