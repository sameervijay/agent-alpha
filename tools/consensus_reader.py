"""
Consensus Estimates Reader
===========================
Reads and parses consensus estimates from Excel files, providing structured
access to revenue, margin, and growth forecasts.

Usage:
    reader = ConsensusReader('NVDA')
    consensus = reader.get_consensus()
    # Returns dict with revenue, segments, margins, growth rates by period
"""

import openpyxl
from pathlib import Path
from typing import Dict, Optional
import config


class ConsensusReader:
    """Read consensus estimates from Excel files."""

    def __init__(self, ticker: str):
        self.ticker = ticker
        self.filepath = self._find_consensus_file()
        self._data = None

    def _find_consensus_file(self) -> Path:
        """Find the consensus estimates file for this ticker."""
        consensus_dir = config._PROJECT_ROOT / 'Consensus estimates'

        # Look for files matching ticker pattern
        candidates = list(consensus_dir.glob(f'{self.ticker}*.xlsx'))
        if not candidates:
            raise FileNotFoundError(
                f"No consensus file found for {self.ticker} in {consensus_dir}"
            )

        # Use most recent file (sort by name, which includes date)
        return sorted(candidates)[-1]

    def _read_excel(self) -> Dict:
        """Read the Excel file and extract structured data."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)

        # NVDA consensus uses ticker as sheet name
        sheet_name = f'{self.ticker}-US'
        if sheet_name not in wb.sheetnames:
            # Fallback to first sheet
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Parse period headers (Row 3) and quarter labels (Row 4)
        periods = self._parse_periods(ws)

        # Extract key metrics
        data = {
            'ticker': self.ticker,
            'source_file': str(self.filepath),
            'periods': periods,
            'revenue': self._extract_metric(ws, 'Sales', 17, periods),
            'segments': {
                'datacenter': self._extract_metric(ws, 'Data Center', 19, periods),
                'gaming': self._extract_metric(ws, 'Gaming', 22, periods),
                'proviz': self._extract_metric(ws, 'Professional Visualization', 23, periods),
                'automotive': self._extract_metric(ws, 'Automotive', 24, periods),
                'oem': self._extract_metric(ws, 'OEM & Other', 25, periods),
            },
            'margins': {
                'gross_income': self._extract_metric(ws, 'Gross Income', 44, periods),
                'operating_income': self._extract_metric(ws, 'Operating Income', 55, periods),
                'ebitda': self._extract_metric(ws, 'EBITDA', 51, periods),
            },
            'opex': {
                'rd': self._extract_metric(ws, 'Research & Development', 50, periods),
                'sga': self._extract_metric(ws, 'SG&A Expense', 47, periods),
            },
            'eps': {
                'gaap': self._extract_metric(ws, 'EPS - GAAP', 7, periods),
                'non_gaap': self._extract_metric(ws, 'EPS - Non GAAP', 11, periods),
            },
        }

        wb.close()
        return data

    def _parse_periods(self, ws) -> Dict[str, Dict]:
        """Parse period headers from rows 3-4."""
        periods = {}

        for col in range(2, ws.max_column + 1):
            period_label = ws.cell(3, col).value
            quarter_label = ws.cell(4, col).value

            if period_label:
                period_str = str(period_label).strip()

                # Determine if it's annual or quarterly
                if quarter_label and str(quarter_label).strip():
                    # Quarterly
                    qtr = str(quarter_label).strip().replace(' ', '')
                    # Convert "Jan '26E" + "Q4" -> "Q4-26"
                    if "'" in period_str:
                        year_part = period_str.split("'")[1].replace('E', '').strip()
                        period_key = f'{qtr}-{year_part}'
                    else:
                        period_key = f'{qtr}-{period_str}'
                else:
                    # Annual - "Jan '26E" -> "FY2026"
                    if "'" in period_str:
                        year_part = period_str.split("'")[1].replace('E', '').strip()
                        # Convert '26 -> 2026
                        if len(year_part) == 2:
                            year_part = '20' + year_part
                        period_key = f'FY{year_part}'
                    else:
                        period_key = period_str

                periods[period_key] = {
                    'col': col,
                    'label': period_str,
                    'quarter': quarter_label if quarter_label else None,
                    'is_annual': not bool(quarter_label),
                }

        return periods

    def _extract_metric(self, ws, label: str, row: int, periods: Dict) -> Dict:
        """Extract a metric row into a dict of {period: value}."""
        values = {}

        for period_key, period_info in periods.items():
            col = period_info['col']
            val = ws.cell(row, col).value

            if val is not None and val != '' and val != '-':
                try:
                    values[period_key] = float(val)
                except (ValueError, TypeError):
                    pass

        return values

    def get_consensus(self) -> Dict:
        """
        Get all consensus data.

        Returns:
            Dict with structure:
            {
                'ticker': str,
                'source_file': str,
                'periods': {...},
                'revenue': {period: value},
                'segments': {segment_name: {period: value}},
                'margins': {...},
                'opex': {...},
                'eps': {...},
            }
        """
        if self._data is None:
            self._data = self._read_excel()
        return self._data

    def get_annual_consensus(self, fiscal_years: list = None) -> Dict:
        """
        Get annual consensus for specific fiscal years.

        Args:
            fiscal_years: List like ['FY2026', 'FY2027', 'FY2028']
                         If None, returns all available annual periods

        Returns:
            Dict filtered to annual periods only
        """
        data = self.get_consensus()

        if fiscal_years is None:
            # Auto-detect annual periods
            fiscal_years = [p for p in data['periods'].keys()
                           if p.startswith('FY') and data['periods'][p]['is_annual']]

        # Filter all metrics to only these periods
        filtered = {
            'ticker': data['ticker'],
            'fiscal_years': fiscal_years,
            'revenue': {fy: data['revenue'].get(fy) for fy in fiscal_years},
            'segments': {},
            'margins': {},
            'opex': {},
            'eps': {},
        }

        for seg, vals in data['segments'].items():
            filtered['segments'][seg] = {fy: vals.get(fy) for fy in fiscal_years}

        for margin, vals in data['margins'].items():
            filtered['margins'][margin] = {fy: vals.get(fy) for fy in fiscal_years}

        for item, vals in data['opex'].items():
            filtered['opex'][item] = {fy: vals.get(fy) for fy in fiscal_years}

        for eps_type, vals in data['eps'].items():
            filtered['eps'][eps_type] = {fy: vals.get(fy) for fy in fiscal_years}

        return filtered

    def print_summary(self):
        """Print a summary of consensus estimates."""
        data = self.get_consensus()

        print(f"\n{'='*80}")
        print(f"  CONSENSUS ESTIMATES: {data['ticker']}")
        print(f"  Source: {Path(data['source_file']).name}")
        print(f"{'='*80}")

        # Get annual periods
        annual_periods = sorted([p for p in data['periods'].keys()
                                if data['periods'][p]['is_annual'] and p.startswith('FY')])

        if not annual_periods:
            print("  No annual forecast periods found")
            return

        # Revenue table
        print(f"\n  Revenue Consensus ($ millions):")
        print(f"  {'Metric':<30} " + "".join(f"{p:>12}" for p in annual_periods))
        print(f"  {'-'*78}")

        # Total revenue
        row = f"  {'Total Revenue':<30} "
        for fy in annual_periods:
            val = data['revenue'].get(fy)
            row += f"{val:>12,.0f}" if val else f"{'N/A':>12}"
        print(row)

        # Segments
        seg_labels = {
            'datacenter': 'Datacenter',
            'gaming': 'Gaming',
            'proviz': 'ProViz',
            'automotive': 'Automotive',
            'oem': 'OEM & Other',
        }

        for seg_key, seg_label in seg_labels.items():
            row = f"  {f'  {seg_label}':<30} "
            for fy in annual_periods:
                val = data['segments'][seg_key].get(fy)
                row += f"{val:>12,.0f}" if val else f"{'N/A':>12}"
            print(row)

        # Y/Y Growth
        print()
        row = f"  {'Y/Y Revenue Growth':<30} "
        for i, fy in enumerate(annual_periods):
            if i > 0:
                curr = data['revenue'].get(fy)
                prev = data['revenue'].get(annual_periods[i-1])
                if curr and prev and prev > 0:
                    growth = (curr / prev - 1)
                    row += f"{growth:>11.1%} "
                else:
                    row += f"{'N/A':>12}"
            else:
                row += f"{'—':>12}"
        print(row)

        # Margins
        print(f"\n  Margin Consensus ($ millions):")
        print(f"  {'Metric':<30} " + "".join(f"{p:>12}" for p in annual_periods))
        print(f"  {'-'*78}")

        for margin_key, margin_label in [('gross_income', 'Gross Income'),
                                          ('operating_income', 'Operating Income'),
                                          ('ebitda', 'EBITDA')]:
            row = f"  {margin_label:<30} "
            for fy in annual_periods:
                val = data['margins'][margin_key].get(fy)
                row += f"{val:>12,.0f}" if val else f"{'N/A':>12}"
            print(row)

        # Margin %
        print()
        row = f"  {'Gross Margin %':<30} "
        for fy in annual_periods:
            rev = data['revenue'].get(fy)
            gm = data['margins']['gross_income'].get(fy)
            if rev and gm and rev > 0:
                row += f"{gm/rev:>11.1%} "
            else:
                row += f"{'N/A':>12}"
        print(row)

        row = f"  {'Operating Margin %':<30} "
        for fy in annual_periods:
            rev = data['revenue'].get(fy)
            oi = data['margins']['operating_income'].get(fy)
            if rev and oi and rev > 0:
                row += f"{oi/rev:>11.1%} "
            else:
                row += f"{'N/A':>12}"
        print(row)

        # EPS
        print(f"\n  EPS Consensus:")
        print(f"  {'Metric':<30} " + "".join(f"{p:>12}" for p in annual_periods))
        print(f"  {'-'*78}")

        for eps_type, eps_label in [('gaap', 'EPS (GAAP)'), ('non_gaap', 'EPS (Non-GAAP)')]:
            row = f"  {eps_label:<30} "
            for fy in annual_periods:
                val = data['eps'][eps_type].get(fy)
                row += f"${val:>11.2f} " if val else f"{'N/A':>12}"
            print(row)

        print()


if __name__ == '__main__':
    import sys
    ticker = sys.argv[1] if len(sys.argv) > 1 else 'NVDA'

    reader = ConsensusReader(ticker)
    reader.print_summary()

    # Example: Get FY2026-2028 annual consensus
    print("\n" + "="*80)
    print("  ANNUAL CONSENSUS (FY2026-2028)")
    print("="*80)

    annual = reader.get_annual_consensus(['FY2026', 'FY2027', 'FY2028'])

    print(f"\nRevenue:")
    for fy in ['FY2026', 'FY2027', 'FY2028']:
        val = annual['revenue'][fy]
        print(f"  {fy}: ${val:,.0f}M" if val else f"  {fy}: N/A")

    print(f"\nDatacenter Segment:")
    for fy in ['FY2026', 'FY2027', 'FY2028']:
        val = annual['segments']['datacenter'][fy]
        print(f"  {fy}: ${val:,.0f}M" if val else f"  {fy}: N/A")
