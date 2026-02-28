#!/usr/bin/env python3
"""
DCF Implementation Guide - Cell Mapping and Access Patterns

This file documents the patterns and methods for extracting key DCF cells
from TSMC, CoreWeave, and ASML analyst models.

Author: Claude Code
Date: February 17, 2026
"""

from openpyxl.utils import get_column_letter, column_index_from_string


class AnalystModelDCFExtractionGuide:
    """
    Reference guide for extracting DCF-critical cells from analyst models.

    Each company has a different row structure, but similar column structure.
    All use standardized Canalyst/AlphaSense templates.
    """

    # =========================================================================
    # SECTION 1: UNIVERSAL CONSTANTS
    # =========================================================================

    HEADER_ROW = 5  # Period labels (FY2025, Q1-26, etc.)
    DATE_ROW = 4    # Date values
    COMPANY_NAME_ROW = 1

    # Column where metric descriptions start
    DESCRIPTION_COL = 'A'
    DESCRIPTION_COL_IDX = 1

    # First data column (after description)
    FIRST_DATA_COL = 'C'
    FIRST_DATA_COL_IDX = 3


    # =========================================================================
    # SECTION 2: COMPANY-SPECIFIC ROW MAPPINGS
    # =========================================================================

    TSMC_ROWS = {
        # Revenue Structure (Rows 6-31)
        'revenue_build_header': 6,
        'smartphone_revenue': 7,
        'smartphone_qq_growth': 8,
        'smartphone_yoy_growth': 9,
        'hpc_revenue': 10,
        'hpc_qq_growth': 11,
        'hpc_yoy_growth': 12,
        'iot_revenue': 13,
        'iot_qq_growth': 14,
        'iot_yoy_growth': 15,
        'automotive_revenue': 16,
        'automotive_qq_growth': 17,
        'automotive_yoy_growth': 18,
        'dce_revenue': 19,
        'dce_qq_growth': 20,
        'dce_yoy_growth': 21,
        'other_revenue': 22,
        'other_qq_growth': 23,
        'other_yoy_growth': 24,
        'total_revenue': 25,
        'total_qq_growth': 27,
        'total_yoy_growth': 28,
        'consolidated_revenue_usd': 30,

        # Profitability (Rows 33-62)
        'opex_section_header': 33,
        'revenue_reference': 34,
        'gross_profit': 36,
        'gross_profit_yoy_growth': 37,
        'gross_margin_pct': 38,
        'consensus_gm_pct': 39,
        'gross_margin_yoy_improvement': 40,
        'rd_expense': 42,
        'rd_yoy_growth': 43,
        'rd_margin_pct': 44,
        'rd_yoy_improvement': 45,
        'ga_margin_pct': 49,
        'ga_yoy_improvement': 50,
        'sm_margin_pct': 54,
        'sm_yoy_improvement': 55,
        'ebit': 59,
        'ebit_yoy_growth': 60,
        'ebit_margin_pct': 61,
        'ebit_yoy_improvement': 62,

        # D&A & Cash Flow (Rows 64-86)
        'da_section_header': 64,
        'depreciation_cogs': 65,
        'depreciation_opex': 66,
        'depreciation_other_opex': 67,
        'total_depreciation': 68,
        'depreciation_rou_assets': 70,
        'depreciation_ppe_calculated': 71,
        'total_depreciation_summary': 72,
        'amortization_cogs': 74,
        'amortization_opex': 75,
        'total_amortization': 76,
        'net_revenue': 80,
        'net_income': 82,
        'eps_wad': 83,
        'eps_adr': 84,
        'capex_bn': 86,
    }

    COREWEAVE_ROWS = {
        # Revenue Structure (Rows 6-10)
        'revenue_build_header': 6,
        'total_revenue': 7,
        'total_qq_growth': 9,
        'total_yoy_growth': 10,

        # EBITDA & Operating Income (Rows 12-31)
        'nongaap_ebitda': 12,
        'ebitda_yoy_growth': 14,
        'nongaap_ebitda_margin': 15,
        'ebitda_margin_yoy_improvement': 17,
        'nongaap_op_income': 19,
        'op_income_yoy_growth': 21,
        'nongaap_op_income_margin': 22,
        'op_income_margin_yoy_improvement': 24,
        'nongaap_net_income': 26,
        'net_income_yoy_growth': 27,
        'nongaap_net_income_margin': 28,
        'net_income_margin_yoy_improvement': 29,
        'nongaap_eps_wad': 31,

        # Operating Expenses (Rows 39-129)
        'opex_section_header': 39,
        'revenue_reference': 40,
        'gaap_gross_profit': 42,
        'gaap_gp_yoy_growth': 43,
        'gaap_gross_margin': 44,  # PRIMARY: Use for DCF
        'gaap_gm_yoy_improvement': 45,
        'sbc_cogs_pct': 48,
        'amortization_intangibles_cogs': 49,
        'nongaap_gross_profit': 51,
        'nongaap_gp_yoy_growth': 52,
        'nongaap_gross_margin': 53,
        'consensus_nongaap_gm': 54,
        'nongaap_gm_yoy_improvement': 55,
        'gaap_rd_expense': 57,
        'gaap_rd_yoy_growth': 58,
        'gaap_rd_margin': 59,  # PRIMARY: Use for DCF
        'gaap_rd_yoy_improvement': 60,
        'sbc_rd': 62,
        'sbc_rd_pct': 63,
        'amortization_intangibles_rd': 64,
        'other_adjustments_rd': 65,
        'nongaap_rd_expense': 66,
        'nongaap_rd_yoy_growth': 67,
        'nongaap_rd_margin': 68,
        'nongaap_rd_yoy_improvement': 69,
        'gaap_sm_margin': 73,  # PRIMARY: Use for DCF
        'gaap_sm_yoy_improvement': 74,
        'sbc_sm_pct': 77,
        'amortization_intangibles_sm': 78,
        'nongaap_sm_margin': 82,
        'nongaap_sm_yoy_improvement': 83,
        'gaap_ga_margin': 87,  # PRIMARY: Use for DCF
        'gaap_ga_yoy_improvement': 88,
        'sbc_ga_pct': 91,
        'amortization_intangibles_ga': 92,
        'nongaap_ga_margin': 96,
        'nongaap_ga_yoy_improvement': 97,
        'gaap_op_income': 99,  # PRIMARY: Use for DCF
        'gaap_op_income_yoy_growth': 100,
        'gaap_op_income_margin': 101,
        'gaap_op_income_margin_yoy_improvement': 102,
        'nongaap_op_income': 104,
        'nongaap_op_income_margin': 107,
        'reported_depreciation': 111,
        'nongaap_ebitda_summary': 113,
        'sbc_rd_summary': 122,
        'depreciation': 127,
        'amortization': 128,
        'total_da': 129,

        # Free Cash Flow (Rows 182+)
        'fcf': 182,
        'ltm_fcf': 183,
        'ltm_revenue': 184,
        'ltm_revenue_yoy_growth': 185,
        'ltm_fcf_margin': 186,
        'capex': 216,

        # Geographic & Customer (Rows 132-172)
        'us_revenue': 133,
        'other_countries_revenue': 134,
        'total_revenue_geo': 135,
        'committed_contracts_revenue_mix': 166,
        'ondemand_revenue_mix': 167,
        'committed_contracts_revenue': 170,
        'ondemand_revenue': 171,
        'total_revenue_by_type': 172,
    }

    ASML_ROWS = {
        # Revenue Structure (Rows 6-67)
        'revenue_build_header': 6,
        'euv_3400c_units': 7,
        'euv_3400b_units': 8,
        'high_na_euv_units': 9,
        'total_revenue': 41,
        'total_yoy_growth': 51,
        'net_deferred_revenue': 66,
        'normalized_revenue': 67,

        # Profitability (Rows 69-101)
        'opex_section_header': 69,
        'revenue_reference': 70,
        'gross_profit': 72,
        'gp_yoy_growth': 73,
        'gross_margin_pct': 74,  # PRIMARY: Use for DCF
        'gm_yoy_improvement': 76,
        'rd_expense': 78,
        'rd_yoy_growth': 79,
        'rd_margin_pct': 80,  # PRIMARY: Use for DCF
        'rd_yoy_improvement': 81,
        'sga_margin_pct': 85,  # PRIMARY: Use for DCF
        'sga_yoy_improvement': 86,
        'ebit': 89,  # PRIMARY: Use for DCF
        'ebit_yoy_growth': 91,
        'ebit_margin_pct': 92,
        'ebit_yoy_improvement': 94,
        'ebitda': 98,
        'ebitda_yoy_growth': 99,
        'ebitda_margin_pct': 100,
        'ebitda_yoy_improvement': 101,
        'add_back_da': 96,

        # D&A (Rows 103-113)
        'da_section_header': 103,
        'depreciation_cogs': 104,
        'depreciation_rd': 105,
        'depreciation_sga': 106,
        'depreciation_ppe': 107,
        'amortization_cogs': 108,
        'amortization_rd': 109,
        'amortization_sga': 110,
        'amortization_intangibles': 111,
        'other_da': 112,
        'total_da': 113,  # PRIMARY: Use for DCF

        # Other Items
        'sbc': 251,
        'nongaap_ebitda': 113,  # Alternative EBITDA row
    }


    # =========================================================================
    # SECTION 3: COLUMN STRUCTURE MAPPING
    # =========================================================================

    @staticmethod
    def get_annual_columns_tsmc_asml():
        """
        TSMC & ASML use same column pattern (81 total columns C through AW).
        Returns mapping of fiscal year to column letter.

        Note: Actual columns depend on model vintage. This shows standard pattern.
        """
        return {
            'FY2024': 'G',
            'FY2025': 'L',
            'FY2026': 'Q',
            'FY2027': 'V',
            'FY2028': 'AA',
            'FY2029': 'AF',
            'FY2030': 'AK',
        }

    @staticmethod
    def get_quarterly_columns_tsmc_asml():
        """TSMC & ASML quarterly columns (for FY2025-FY2027)."""
        return {
            'Q1-2025': 'H',
            'Q2-2025': 'I',
            'Q3-2025': 'J',
            'Q4-2025': 'K',
            'Q1-2026': 'M',
            'Q2-2026': 'N',
            'Q3-2026': 'O',
            'Q4-2026': 'P',
            'Q1-2027': 'R',
            'Q2-2027': 'S',
            'Q3-2027': 'T',
            'Q4-2027': 'U',
            'Q1-2028': 'W',
            'Q2-2028': 'X',
            'Q3-2028': 'Y',
            'Q4-2028': 'Z',
        }

    @staticmethod
    def get_annual_columns_coreweave():
        """
        CoreWeave uses compressed column pattern (48 total columns C through AU).
        Returns mapping of fiscal year to column letter.
        """
        return {
            'FY2024': 'N',
            'FY2025': 'S',
            'FY2026': 'X',
            'FY2027': 'AC',
            'FY2028': 'AH',
        }

    @staticmethod
    def get_quarterly_columns_coreweave():
        """CoreWeave quarterly columns."""
        return {
            'Q1-2025': 'O',
            'Q2-2025': 'P',
            'Q3-2025': 'Q',
            'Q4-2025': 'R',
            'Q1-2026': 'T',
            'Q2-2026': 'U',
            'Q3-2026': 'V',
            'Q4-2026': 'W',
            'Q1-2027': 'Y',
            'Q2-2027': 'Z',
            'Q3-2027': 'AA',
            'Q4-2027': 'AB',
            'Q1-2028': 'AD',
            'Q2-2028': 'AE',
            'Q3-2028': 'AF',
            'Q4-2028': 'AG',
        }


    # =========================================================================
    # SECTION 4: EXTRACTION METHODS (PSEUDOCODE)
    # =========================================================================

    @staticmethod
    def get_period_column_dynamic(ws, target_period):
        """
        Dynamically find the column letter for a given period.

        Args:
            ws: openpyxl Worksheet object
            target_period: str like "FY2025" or "Q1-26"

        Returns:
            str: Column letter (e.g., "L") or None if not found

        Example:
            >>> col = get_period_column_dynamic(ws, "FY2026")
            >>> # Returns "Q"
        """
        for col_idx in range(3, ws.max_column + 1):
            label = ws.cell(5, col_idx).value  # Row 5 = period labels
            if label and target_period in str(label):
                return get_column_letter(col_idx)
        return None

    @staticmethod
    def extract_metric(ws, row, col_letter):
        """
        Extract a single metric value from worksheet.

        Args:
            ws: openpyxl Worksheet
            row: int, row number
            col_letter: str, column letter (e.g., "L")

        Returns:
            float or str, the cell value
        """
        cell = ws[f"{col_letter}{row}"]
        return cell.value

    @staticmethod
    def extract_revenue_structure_tsmc(ws, col_letter):
        """
        Extract all 5 TSMC revenue segments for a given period.

        Returns:
            dict: {
                'smartphone': float,
                'hpc': float,
                'iot': float,
                'automotive': float,
                'dce': float,
                'other': float,
                'total': float,
            }
        """
        rows = AnalystModelDCFExtractionGuide.TSMC_ROWS
        return {
            'smartphone': ws[f"{col_letter}{rows['smartphone_revenue']}"].value,
            'hpc': ws[f"{col_letter}{rows['hpc_revenue']}"].value,
            'iot': ws[f"{col_letter}{rows['iot_revenue']}"].value,
            'automotive': ws[f"{col_letter}{rows['automotive_revenue']}"].value,
            'dce': ws[f"{col_letter}{rows['dce_revenue']}"].value,
            'other': ws[f"{col_letter}{rows['other_revenue']}"].value,
            'total': ws[f"{col_letter}{rows['total_revenue']}"].value,
        }

    @staticmethod
    def extract_margins_tsmc(ws, col_letter):
        """
        Extract all margin metrics for TSMC.

        Returns:
            dict: All margin percentages and bps improvements
        """
        rows = AnalystModelDCFExtractionGuide.TSMC_ROWS
        return {
            'gross_margin_pct': ws[f"{col_letter}{rows['gross_margin_pct']}"].value,
            'rd_margin_pct': ws[f"{col_letter}{rows['rd_margin_pct']}"].value,
            'ga_margin_pct': ws[f"{col_letter}{rows['ga_margin_pct']}"].value,
            'sm_margin_pct': ws[f"{col_letter}{rows['sm_margin_pct']}"].value,
            'ebit_margin_pct': ws[f"{col_letter}{rows['ebit_margin_pct']}"].value,
            'gross_margin_improvement_bps': ws[f"{col_letter}{rows['gross_margin_yoy_improvement']}"].value,
            'rd_margin_improvement_bps': ws[f"{col_letter}{rows['rd_yoy_improvement']}"].value,
            'ga_margin_improvement_bps': ws[f"{col_letter}{rows['ga_yoy_improvement']}"].value,
            'sm_margin_improvement_bps': ws[f"{col_letter}{rows['sm_yoy_improvement']}"].value,
        }

    @staticmethod
    def extract_cash_flow_tsmc(ws, col_letter):
        """Extract D&A and CapEx for TSMC."""
        rows = AnalystModelDCFExtractionGuide.TSMC_ROWS
        capex_bn = ws[f"{col_letter}{rows['capex_bn']}"].value
        capex_mm = capex_bn * 1000 if capex_bn else None  # Convert to millions

        return {
            'depreciation_mm': ws[f"{col_letter}{rows['total_depreciation']}"].value,
            'amortization_mm': ws[f"{col_letter}{rows['total_amortization']}"].value,
            'da_mm': (
                (ws[f"{col_letter}{rows['total_depreciation']}"].value or 0) +
                (ws[f"{col_letter}{rows['total_amortization']}"].value or 0)
            ),
            'capex_mm': capex_mm,
        }

    @staticmethod
    def extract_metrics_coreweave(ws, col_letter):
        """
        Extract key metrics for CoreWeave (using GAAP basis).

        NOTE: CoreWeave reports both GAAP and Non-GAAP. For DCF use GAAP.
        """
        rows = AnalystModelDCFExtractionGuide.COREWEAVE_ROWS
        return {
            # Revenue
            'revenue_mm': ws[f"{col_letter}{rows['total_revenue']}"].value,
            'revenue_yoy_growth': ws[f"{col_letter}{rows['total_yoy_growth']}"].value,

            # GAAP Margins (PRIMARY)
            'gross_margin_gaap': ws[f"{col_letter}{rows['gaap_gross_margin']}"].value,
            'rd_margin_gaap': ws[f"{col_letter}{rows['gaap_rd_margin']}"].value,
            'sm_margin_gaap': ws[f"{col_letter}{rows['gaap_sm_margin']}"].value,
            'ga_margin_gaap': ws[f"{col_letter}{rows['gaap_ga_margin']}"].value,
            'op_income_gaap': ws[f"{col_letter}{rows['gaap_op_income']}"].value,
            'op_income_margin_gaap': ws[f"{col_letter}{rows['gaap_op_income_margin']}"].value,

            # D&A (Total)
            'da_mm': ws[f"{col_letter}{rows['total_da']}"].value,

            # CapEx
            'capex_mm': ws[f"{col_letter}{rows['capex']}"].value,

            # Free Cash Flow
            'fcf_mm': ws[f"{col_letter}{rows['fcf']}"].value,
        }

    @staticmethod
    def extract_metrics_asml(ws, col_letter):
        """Extract key metrics for ASML."""
        rows = AnalystModelDCFExtractionGuide.ASML_ROWS
        return {
            # Revenue
            'revenue_mm': ws[f"{col_letter}{rows['total_revenue']}"].value,
            'revenue_yoy_growth': ws[f"{col_letter}{rows['total_yoy_growth']}"].value,
            'normalized_revenue_mm': ws[f"{col_letter}{rows['normalized_revenue']}"].value,

            # Profitability
            'gross_profit_mm': ws[f"{col_letter}{rows['gross_profit']}"].value,
            'gross_margin_pct': ws[f"{col_letter}{rows['gross_margin_pct']}"].value,
            'rd_margin_pct': ws[f"{col_letter}{rows['rd_margin_pct']}"].value,
            'sga_margin_pct': ws[f"{col_letter}{rows['sga_margin_pct']}"].value,
            'ebit_mm': ws[f"{col_letter}{rows['ebit']}"].value,
            'ebit_margin_pct': ws[f"{col_letter}{rows['ebit_margin_pct']}"].value,
            'ebitda_mm': ws[f"{col_letter}{rows['ebitda']}"].value,
            'ebitda_margin_pct': ws[f"{col_letter}{rows['ebitda_margin_pct']}"].value,

            # D&A
            'da_mm': ws[f"{col_letter}{rows['total_da']}"].value,

            # Unit Metrics (Product-based model)
            'euv_3400c_units': ws[f"{col_letter}{rows['euv_3400c_units']}"].value,
            'euv_3400b_units': ws[f"{col_letter}{rows['euv_3400b_units']}"].value,
        }


    # =========================================================================
    # SECTION 5: DCF BUILDING BLOCKS
    # =========================================================================

    @staticmethod
    def calculate_ufcf(revenue_mm, ebit_margin_pct, tax_rate_pct,
                       da_mm, capex_mm, nwc_change_mm):
        """
        Calculate Unlevered Free Cash Flow (UFCF).

        UFCF = EBIT * (1 - Tax Rate) + D&A - CapEx - Change in NWC
            = NOPAT + D&A - CapEx - Change in NWC

        Where:
            NOPAT = EBIT * (1 - Tax Rate) = Operating profit after tax
        """
        ebit_mm = revenue_mm * (ebit_margin_pct / 100)
        nopat_mm = ebit_mm * (1 - tax_rate_pct / 100)
        ufcf_mm = nopat_mm + da_mm - capex_mm - nwc_change_mm
        return {
            'ebit_mm': ebit_mm,
            'nopat_mm': nopat_mm,
            'ufcf_mm': ufcf_mm,
        }

    @staticmethod
    def cascade_margins_with_improvement(baseline_margin_pct, yoy_improvement_bps,
                                        num_periods=5):
        """
        Cascade margin improvements period-over-period.

        Used for scenario analysis: as PM agent adjusts margin improvement bps,
        propagate through forecast.

        Args:
            baseline_margin_pct: Starting margin (e.g., 50.0)
            yoy_improvement_bps: Annual improvement in basis points (e.g., 200)
            num_periods: Number of years to project

        Returns:
            list: Margin % for each period
        """
        margins = [baseline_margin_pct]
        for _ in range(num_periods - 1):
            new_margin = margins[-1] + (yoy_improvement_bps / 10000)
            margins.append(new_margin)
        return margins


    # =========================================================================
    # SECTION 6: VALIDATION & ERROR HANDLING
    # =========================================================================

    @staticmethod
    def validate_metric(value, metric_name, min_val=None, max_val=None):
        """
        Validate extracted metric for reasonableness.

        Returns:
            tuple: (is_valid: bool, error_msg: str or None)
        """
        if value is None:
            return False, f"{metric_name}: Value is None"

        if not isinstance(value, (int, float)):
            return False, f"{metric_name}: Not numeric, got {type(value)}"

        if min_val is not None and value < min_val:
            return False, f"{metric_name}: {value} < min {min_val}"

        if max_val is not None and value > max_val:
            return False, f"{metric_name}: {value} > max {max_val}"

        return True, None

    @staticmethod
    def validate_revenue_buildup(total_revenue, segment_revenues_dict):
        """
        Validate that segments sum to total revenue (within rounding).

        Returns:
            tuple: (matches: bool, variance_pct: float)
        """
        segment_sum = sum(segment_revenues_dict.values())
        if total_revenue == 0:
            return False, None

        variance_pct = abs((segment_sum - total_revenue) / total_revenue) * 100
        # Allow 0.5% variance due to rounding
        matches = variance_pct < 0.5
        return matches, variance_pct


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == "__main__":
    """
    Example: Extract and validate TSMC DCF metrics for FY2026.

    Assuming openpyxl is available and file is loaded as wb.
    """

    # import openpyxl
    #
    # # Load workbook
    # wb = openpyxl.load_workbook('NVIDIA NVDA US.xlsx')
    # ws = wb['Model']
    #
    # guide = AnalystModelDCFExtractionGuide()
    #
    # # Find FY2026 column dynamically
    # fy2026_col = guide.get_period_column_dynamic(ws, "FY2026")
    # print(f"FY2026 column: {fy2026_col}")  # Should print "Q"
    #
    # # Extract revenue structure
    # revenue = guide.extract_revenue_structure_tsmc(ws, fy2026_col)
    # print(f"FY2026 Total Revenue: ${revenue['total']}mm")
    #
    # # Validate revenue buildup
    # segment_revenues = {k: v for k, v in revenue.items() if k != 'total'}
    # is_valid, variance = guide.validate_revenue_buildup(
    #     revenue['total'],
    #     segment_revenues
    # )
    # print(f"Revenue segments match total: {is_valid} (variance: {variance}%)")
    #
    # # Extract margins
    # margins = guide.extract_margins_tsmc(ws, fy2026_col)
    # print(f"FY2026 Gross Margin: {margins['gross_margin_pct']}%")
    # print(f"FY2026 EBIT Margin: {margins['ebit_margin_pct']}%")
    #
    # # Extract cash flow
    # cf = guide.extract_cash_flow_tsmc(ws, fy2026_col)
    # print(f"FY2026 Total D&A: ${cf['da_mm']}mm")
    # print(f"FY2026 CapEx: ${cf['capex_mm']}mm")

    print("DCF Implementation Guide - Ready for use")
    print(f"Companies supported: TSMC, CoreWeave, ASML")
    print(f"Methods: Dynamic period lookup, metric extraction, validation")
