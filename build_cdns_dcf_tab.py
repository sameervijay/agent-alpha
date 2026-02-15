"""
build_cdns_dcf_tab.py
=====================
Builds the full DCF tab in the Cadence analyst model Excel file.
Includes: Revenue build by segment (3 segments), operating margin drivers (4),
UFCF build, discounting, valuation, equity bridge, and memo items.

Usage:
    cd agent-alpha
    python build_cdns_dcf_tab.py

Color coding: Blue = hardcoded inputs, Green = cross-tab refs, Black = local formulas
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

wb = openpyxl.load_workbook('Cadence Design CDNS US.xlsx')

if 'DCF' in wb.sheetnames:
    idx = wb.sheetnames.index('DCF')
    del wb['DCF']
else:
    idx = 1
ws = wb.create_sheet('DCF', idx)

# ── Column mappings ──────────────────────────────────────────
MC = {  # Model tab columns
    'FY2025': 'BS', 'FY2026': 'BX', 'FY2027': 'BY',
    'FY2028': 'BZ', 'FY2029': 'CA',
}
DC = {  # DCF tab columns
    'FY2025': 'F', 'FY2026': 'G', 'FY2027': 'H',
    'FY2028': 'I', 'FY2029': 'J',
}
ALL = ['FY2025', 'FY2026', 'FY2027', 'FY2028', 'FY2029']
PROJ = ['FY2026', 'FY2027', 'FY2028', 'FY2029']

# ── Colors & Fonts ───────────────────────────────────────────
BLUE   = '0000FF'
GREEN  = '007F00'
BLACK  = '000000'

def mkfont(color=BLACK, bold=False, size=None, italic=False, underline=None):
    kw = {'color': color, 'bold': bold}
    if size: kw['size'] = size
    if italic: kw['italic'] = italic
    if underline: kw['underline'] = underline
    return Font(**kw)

f_title    = mkfont(BLACK, True, 14)
f_sub      = mkfont(BLACK, True, 11)
f_section  = mkfont(BLACK, True, 11, underline='single')
f_bold     = mkfont(BLACK, True)
f_normal   = mkfont(BLACK)
f_ital     = mkfont(BLACK, italic=True, size=9)
f_note     = mkfont('666666', italic=True, size=9)
f_blue     = mkfont(BLUE)
f_blue_b   = mkfont(BLUE, True)
f_green    = mkfont(GREEN)
f_green_b  = mkfont(GREEN, True)
f_black_b  = mkfont(BLACK, True)
f_result   = mkfont(BLACK, True, 12)
f_upside   = mkfont('0000CC', True, 11)
f_assum_h  = mkfont(BLACK, True, 12, underline='single')

# Number formats
pct   = '0.0%'
num   = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
dlr   = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
dec3  = '_(* #,##0.000_);_(* \\(#,##0.000\\);_(* "-"??_);_(@_)'
price = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
multx = '#,##0.0"x"'
bps_f = '#,##0'
ctr   = Alignment(horizontal='center')

thin = Side(style='thin')
bot_border  = Border(bottom=thin)
top_dbl     = Border(top=thin, bottom=Side(style='double'))

# Column widths
for c, w in {'A': 2, 'B': 44, 'C': 5, 'D': 2, 'E': 2, 'F': 16, 'G': 16,
             'H': 16, 'I': 16, 'J': 16, 'K': 3, 'L': 30, 'M': 2, 'N': 2,
             'O': 12, 'P': 2, 'Q': 38}.items():
    ws.column_dimensions[c].width = w

# ── Helper ───────────────────────────────────────────────────
def cell(addr, val, font=None, fmt=None, align=None, border=None):
    ws[addr] = val
    if font:   ws[addr].font = font
    if fmt:    ws[addr].number_format = fmt
    if align:  ws[addr].alignment = align
    if border: ws[addr].border = border

def model_ref(fy, row):
    """Return Model!{col}{row} reference string."""
    return f"=Model!{MC[fy]}{row}"

# ══════════════════════════════════════════════════════════════
# TITLE & HEADERS
# ══════════════════════════════════════════════════════════════
cell('B1', 'Cadence Design Systems', f_title)
cell('B2', 'Discounted Cash Flow Valuation Analysis', f_sub)

cell('F5', 'Historical', mkfont(BLACK, True, 10), align=ctr)
cell('G5', 'Projected',  mkfont(BLACK, True, 10), align=ctr)
cell('L5', 'Assumptions', f_assum_h)
cell('Q5', 'Notes', f_assum_h)

cell('B7', '($ in millions)', f_ital)
for fy in ALL:
    cell(f'{DC[fy]}7', fy, f_bold, align=ctr, border=bot_border)

# ══════════════════════════════════════════════════════════════
# ASSUMPTIONS PANEL (L6:Q16)
# ══════════════════════════════════════════════════════════════
assumptions = [
    (6,  'Risk Free Rate:',             0.0425,  '0.00%', '10-Year US Treasury yield'),
    (7,  'Market Risk Premium:',        0.055,   '0.00%', 'Damodaran equity risk premium'),
    (8,  'Unlevered Beta:',             1.15,    '0.00',  'EDA software industry avg'),
    (9,  'Projected Equity to Value:',  0.92,    '0.0%',  'CDNS moderate leverage'),
    (10, 'Levered Beta:',              '=(1+(1-O12)*(1-O9)/O9)*O8', '0.00', 'Hamada equation'),
    (11, 'Cost of Equity:',            '=O6+O7*O10', '0.00%', 'CAPM: Rf + B(Rm-Rf)'),
    (12, 'Tax Rate:',                   0.18,    '0%',    'Long-term effective tax rate'),
    (13, 'Cost of Debt:',              0.04,    '0.00%', 'Based on outstanding senior notes'),
    (14, 'WACC:',                      '=O11*O9+O13*(1-O12)*(1-O9)', '0.0%',
          'Rd*(1-T)*(D/V) + Re*(E/V)'),
    (15, 'Terminal Growth Rate:',       0.03,    '0.00%', 'Long-term nominal GDP growth'),
    (16, 'Implied TV/EBITDA:',         None,     multx,   'Terminal value / FY2029 EBITDA'),
]
for r, label, val, fmt, note in assumptions:
    cell(f'L{r}', label, f_bold)
    if val is not None:
        is_formula = isinstance(val, str) and val.startswith('=')
        cell(f'O{r}', val, f_blue if not is_formula else f_normal, fmt=fmt)
    cell(f'Q{r}', note, f_note)

# ══════════════════════════════════════════════════════════════
# REVENUE BUILD (Rows 9–21)
# ══════════════════════════════════════════════════════════════
cell('B9', 'Revenue Build by Segment', f_section)

# Segments: (label, revenue_row, growth_row, dcf_row_rev, dcf_row_growth)
segments = [
    ('Core EDA Revenue',                        13, 14, 10, 11),
    ('System Interconnect & Analysis Revenue',  16, 17, 12, 13),
    ('IP Revenue',                              18, 19, 14, 15),
]

for label, mrev, mgrow, dr, dg in segments:
    cell(f'B{dr}', label, f_normal)
    cell(f'B{dg}', '  Y/Y Growth', f_normal)
    for fy in ALL:
        d = DC[fy]
        cell(f'{d}{dr}', model_ref(fy, mrev), f_green, fmt=num)
        cell(f'{d}{dg}', model_ref(fy, mgrow), f_green, fmt=pct)

# Row 16: Total Revenue
cell('B16', 'Total Revenue', f_green_b)
for fy in ALL:
    cell(f'{DC[fy]}16', model_ref(fy, 20), f_green_b, fmt=num, border=bot_border)

# Row 17: Y/Y Total Revenue Growth (computed locally)
cell('B17', 'Y/Y Total Revenue Growth', f_black_b)
for fy in PROJ:
    d = DC[fy]; p = DC[ALL[ALL.index(fy)-1]]
    cell(f'{d}17', f'={d}16/{p}16-1', f_normal, fmt=pct)

# ══════════════════════════════════════════════════════════════
# MARGIN / OPEX DRIVERS (Rows 19–30)
# ══════════════════════════════════════════════════════════════
cell('B19', 'Operating Margin Drivers', f_section)

# (label, model_row, dcf_row, fmt, font_override)
margin_lines = [
    ('Non-GAAP Gross Margin %',           47,  20, pct,   f_green),
    ('  Gross Margin Improvement (bps)',   49,  21, bps_f, f_green),
    ('Non-GAAP R&D Margin %',             61,  22, pct,   f_green),
    ('  R&D Margin Improvement (bps)',     62,  23, bps_f, f_green),
    ('Non-GAAP G&A Margin %',             74,  24, pct,   f_green),
    ('  G&A Margin Improvement (bps)',     75,  25, bps_f, f_green),
    ('Non-GAAP S&M Margin %',             87,  26, pct,   f_green),
    ('  S&M Margin Improvement (bps)',     88,  27, bps_f, f_green),
    ('Non-GAAP EBIT Margin %',           102,  28, pct,   f_green_b),
]

for label, mrow, drow, fmtv, fnt in margin_lines:
    cell(f'B{drow}', label, f_normal if fnt != f_green_b else f_bold)
    for fy in ALL:
        cell(f'{DC[fy]}{drow}', model_ref(fy, mrow), fnt, fmt=fmtv)

# Row 29: Non-GAAP EBIT
cell('B29', 'Non-GAAP Operating Income (EBIT)', f_green_b)
for fy in ALL:
    cell(f'{DC[fy]}29', model_ref(fy, 99), f_green_b, fmt=num, border=bot_border)

# ══════════════════════════════════════════════════════════════
# UFCF BUILD (Rows 31–38)
# ══════════════════════════════════════════════════════════════
cell('B31', 'Unlevered Free Cash Flow Build', f_section)

# Row 32: EBIT (local ref to row 29)
cell('B32', 'Non-GAAP EBIT', f_bold)
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}32', f'={d}29', f_normal, fmt=num)

# Row 33: Less Taxes
cell('B33', 'Less: Taxes on EBIT')
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}33', f'=-{d}32*$O$12', f_normal, fmt=num)

# Row 34: NOPAT
cell('B34', 'NOPAT', f_bold)
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}34', f'={d}32+{d}33', f_normal, fmt=num)

# Row 35: D&A
cell('B35', 'Plus: Depreciation & Amortization')
for fy in PROJ:
    cell(f'{DC[fy]}35', model_ref(fy, 170), f_green, fmt=num)

# Row 36: CapEx
cell('B36', 'Less: Capital Expenditures')
for fy in PROJ:
    cell(f'{DC[fy]}36', f"=-Model!{MC[fy]}938", f_green, fmt=num)

# Row 37: Change in Working Capital (direct from model, no NWC detail)
cell('B37', 'Less: Increase in Net Working Capital')
for fy in PROJ:
    cell(f'{DC[fy]}37', f"=-Model!{MC[fy]}933", f_green, fmt=num)

# Row 38: UFCF
cell('B38', 'Unlevered Free Cash Flow', f_black_b, border=top_dbl)
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}38', f'=SUM({d}34:{d}37)', f_black_b, fmt=num, border=top_dbl)

# ══════════════════════════════════════════════════════════════
# DISCOUNTING (Rows 40–41)
# ══════════════════════════════════════════════════════════════
cell('B40', 'Times: Discount Factor')
for i, fy in enumerate(PROJ, 1):
    cell(f'{DC[fy]}40', f'=1/(1+$O$14)^{i}', f_normal, fmt=dec3)

cell('B41', 'Discounted Cash Flows')
for fy in PROJ:
    d = DC[fy]
    cell(f'{d}41', f'={d}38*{d}40', f_normal, fmt=num)

# ══════════════════════════════════════════════════════════════
# VALUATION (Rows 43–47)
# ══════════════════════════════════════════════════════════════
cell('B43', 'Sum of Discounted Cash Flows', f_bold)
cell('F43', '=SUM(G41:J41)', f_normal, fmt=dlr)

cell('B44', 'Terminal Value:', f_bold)
cell('F44', '=(J38*(1+$O$15))/($O$14-$O$15)', f_normal, fmt=num)

cell('B45', 'Times: Discount Factor')
cell('F45', '=J40', f_normal, fmt=dec3)

cell('B46', 'PV of Terminal Value')
cell('F46', '=F44*F45', f_normal, fmt=dlr)

cell('C47', 'Enterprise Value', f_result)
cell('F47', '=F43+F46', f_normal, fmt=dlr, border=top_dbl)

# Update implied TV/EBITDA in assumptions panel
# Use EBIT + D&A as proxy for EBITDA (row 99 + row 170)
cell('O16', f"=F44/(Model!{MC['FY2029']}99+Model!{MC['FY2029']}170)", f_green, fmt=multx)

# ══════════════════════════════════════════════════════════════
# EQUITY BRIDGE (Rows 49–59)
# ══════════════════════════════════════════════════════════════
cell('B49', 'Equity Bridge:', f_section)

cell('B50', 'Enterprise Value')
cell('F50', '=F47', f_normal, fmt=dlr)

cell('B51', 'Less: Total Debt')
cell('F51',
     f"=-(Model!{MC['FY2025']}964+IF(ISNUMBER(Model!{MC['FY2025']}963),Model!{MC['FY2025']}963,0))",
     f_green, fmt=dlr)

cell('B52', 'Plus: Cash & Cash Equivalents')
cell('F52', model_ref('FY2025', 962), f_green, fmt=dlr)

cell('C53', 'Equity Value', f_result)
cell('F53', '=SUM(F50:F52)', f_normal, fmt=dlr, border=top_dbl)

cell('B54', 'Diluted Shares Outstanding (mm)')
cell('F54', model_ref('FY2025', 521), f_green, fmt=num)

cell('C55', 'Implied Share Price ($/share)', f_result)
cell('F55', '=F53/F54', f_normal, fmt=price, border=top_dbl)

cell('B57', 'Current Stock Price ($/share)', f_bold)
cell('F57', "='Front Page'!H20", f_green, fmt=price)

cell('B58', 'Upside / (Downside)', f_upside)
cell('F58', '=F55/F57-1', f_upside, fmt=pct)

# ══════════════════════════════════════════════════════════════
# MEMO (Rows 61–64)
# ══════════════════════════════════════════════════════════════
cell('B61', 'Memo:', f_section)

cell('B62', 'Stock-Based Compensation (excluded from UFCF)')
for fy in PROJ:
    cell(f'{DC[fy]}62', model_ref(fy, 178), f_green, fmt=num)

cell('B63', 'UFCF Growth Rate')
for i, fy in enumerate(PROJ):
    if i > 0:
        d = DC[fy]; p = DC[PROJ[i-1]]
        cell(f'{d}63', f'={d}38/{p}38-1', f_normal, fmt=pct)

# ── Save ─────────────────────────────────────────────────────
wb.save('Cadence Design CDNS US.xlsx')
print('DCF tab rebuilt successfully!')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
